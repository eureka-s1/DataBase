from __future__ import annotations

import argparse
import hashlib
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import xlrd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.db import db_session
from app.services.common import to_float, to_int
from app.services.customers import get_or_create_customer_id, resolve_customer_id, upsert_alias
from app.services.inbound import create_inbound_item

DEFAULT_DATA_ROOT = ROOT / "2026data"
EXCLUDED_TOP_DIRS = {"每日收货清单", "装柜清单", "店面对账"}

HEADER_MAP = {
    "CUSTOMER NAME": "customer_name",
    "SHOP NO": "shop_no",
    "位置": "position_or_tel",
    "TEL": "position_or_tel",
    "ITEM NO": "item_no",
    "DESCRIPTION": "item_name_cn",
    "DESCRPETION": "item_name_cn",
    "ITEM NAME": "item_name_cn",
    "ITEMNAME": "item_name_cn",
    "品名": "item_name_cn",
    "材质": "material",
    "CTNS": "carton_count",
    "CTN": "carton_count",
    "CNS": "carton_count",
    "QTY": "qty",
    "PRICE": "unit_price",
    "T.PRICE": "total_price",
    "T.T": "total_price",
    "AMOUNT": "total_price",
    "定金": "deposit_hint",
    "DEP": "deposit_hint",
    "DEPOZIT": "deposit_hint",
    "CBM": "cbm_calculated",
}

PHONE_RE = re.compile(r"^\d{7,15}$")
# Some source files contain typo "FEIGHT"; accept both FEIGHT/FREIGHT.
FREIGHT_RE = re.compile(r"CBM\s*F(?:R)?EIGHT", re.I)
# Container no variants in source files:
# - common: 4 letters + 7 digits (e.g. ONEU6363669)
# - observed: 4 letters + 6 digits (e.g. TRLU711969)
# - observed: 5 letters + 6/7 digits in some sheets
CONTAINER_NO_RE = re.compile(r"^[A-Z]{4,5}\d{6,7}$", re.I)
EN_ALIAS_RE = re.compile(r"^[A-Za-z][A-Za-z ]{0,79}$")


@dataclass
class ParsedItem:
    customer_name: str
    shop_no: str | None
    position_or_tel: str | None
    item_no: str | None
    item_name_cn: str
    material: str | None
    carton_count: int
    qty: int
    unit_price: float
    total_price: float
    deposit_hint: float
    cbm_calculated: float
    length_cm: float
    width_cm: float
    height_cm: float
    inbound_date_hint: str | None
    source_file: str
    source_sheet: str
    source_row_no: int


def _norm_header(value) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def _to_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_ymd(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        # Excel serial date (1900 date system)
        fv = float(value)
        if 30000 <= fv <= 70000:
            d = datetime(1899, 12, 30) + timedelta(days=int(fv))
            return d.strftime("%Y-%m-%d")
    s = _to_text(value)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    m = re.search(r"(20\d{2})[./-](\d{1,2})[./-](\d{1,2})", s)
    if m:
        y, mm, dd = m.groups()
        return f"{int(y):04d}-{int(mm):02d}-{int(dd):02d}"
    return None


def _first_col_phone_token(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        iv = int(v)
        if abs(float(v) - iv) < 1e-9 and iv > 0:
            return str(iv)
    s = str(v).strip().replace(" ", "")
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s


def _is_skip_item_token(value: str) -> bool:
    v = (value or "").strip().upper()
    if not v:
        return True
    bad_exact = {"品名", "ITEM NAME", "ITEM NO", "SHOP NO", "TEL", "TOTAL", "BALANCE", "REMAIN"}
    if v in bad_exact:
        return True
    bad_contains = ("CBM FREIGHT", "SEND TO ME", "PAID", "COMMISSION")
    if any(x in v for x in bad_contains):
        return True
    return bool(re.match(r"^\d+\s*CTNS?$", v))


def _normalize_name_like(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def _read_book(path: Path, max_cols: int = 40) -> list[tuple[str, list[list]]]:
    suffix = path.suffix.lower()
    out: list[tuple[str, list[list]]] = []
    if suffix == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows: list[list] = []
            for r in ws.iter_rows(values_only=True, max_col=max_cols):
                rows.append(list(r))
            out.append((sname, rows))
        wb.close()
        return out
    if suffix == ".xls":
        wb = xlrd.open_workbook(path)
        for i in range(wb.nsheets):
            sh = wb.sheet_by_index(i)
            rows = [sh.row_values(r, 0, max_cols) for r in range(sh.nrows)]
            out.append((sh.name or f"Sheet{i + 1}", rows))
        return out
    raise ValueError("unsupported file type")


def _sheet_has_content(rows: list[list]) -> bool:
    # Business rule:
    # if first row is empty, treat this sheet as empty directly.
    if not rows:
        return False
    first = rows[0] if rows else []
    if not any(_to_text(v) for v in first):
        return False
    for row in rows:
        for v in row:
            if _to_text(v):
                return True
    return False


def _pick_last_non_empty_sheet(book: list[tuple[str, list[list]]]) -> tuple[str, list[list]] | None:
    for sname, rows in reversed(book):
        if _sheet_has_content(rows):
            return sname, rows
    return None


def _sheet_year_from_name(sheet_name: str) -> int | None:
    m = re.search(r"(20\d{2})", str(sheet_name or ""))
    if not m:
        return None
    y = int(m.group(1))
    if 2000 <= y <= 2100:
        return y
    return None


def _looks_like_english_alias_token(value: str) -> bool:
    s = str(value or "").strip()
    if not s:
        return False
    # Alias detection currently only supports plain English aliases.
    if not EN_ALIAS_RE.fullmatch(s):
        return False
    u = s.upper()
    bad_exact = {
        "CUSTOMER", "CUSTOMER NAME", "SHOP NO", "TEL", "ITEM NO", "DESCRIPTION",
        "ITEM NAME", "MATERIAL", "CTNS", "QTY", "PRICE", "T PRICE", "CBM",
        "DATE", "AUTO", "ORDER",
    }
    if u in bad_exact:
        return False
    return True


def _collect_recent_sheet_aliases(book: list[tuple[str, list[list]]], customer_name: str, years: int = 2) -> tuple[list[str], Counter]:
    stats = Counter()
    out: set[str] = set()
    current_year = datetime.now().year
    min_year = current_year - max(1, years) + 1
    customer_norm = _normalize_name_like(customer_name)

    for sname, rows in book:
        y = _sheet_year_from_name(sname)
        if y is None:
            stats["alias_sheets_skipped_unparsable_year"] += 1
            continue
        if y < min_year or y > current_year:
            stats["alias_sheets_skipped_outside_window"] += 1
            continue
        stats["alias_sheets_scanned"] += 1
        for row in rows[:2000]:
            if not row:
                continue
            first = row[0] if len(row) > 0 else None
            text = _to_text(first)
            if not text:
                continue
            phone_token = _first_col_phone_token(first)
            if phone_token and PHONE_RE.fullmatch(phone_token):
                continue
            if not _looks_like_english_alias_token(text):
                continue
            alias_norm = _normalize_name_like(text)
            if not alias_norm or alias_norm == customer_norm:
                continue
            out.add(text.strip())

    stats["aliases_detected"] = len(out)
    return sorted(out), stats


def _detect_header(rows: list[list], max_scan: int = 120) -> tuple[int, dict[int, str]]:
    keys = {k.replace(" ", "").upper(): v for k, v in HEADER_MAP.items()}
    for i, row in enumerate(rows[:max_scan]):
        mapping: dict[int, str] = {}
        for idx, val in enumerate(row):
            kk = _norm_header(val)
            if kk in keys:
                mapping[idx] = keys[kk]
        has_item = ("item_name_cn" in mapping.values()) or ("item_no" in mapping.values())
        if len(mapping) >= 3 and has_item:
            return i, mapping
    raise ValueError("unable to detect header row")


def _last_freight_marker_idx(rows: list[list]) -> int:
    last = -1
    for i, row in enumerate(rows):
        if _is_freight_marker_row(row):
            last = i
    return last


def _is_container_no_token(value: str) -> bool:
    token = (value or "").strip().upper().replace(" ", "")
    if not token:
        return False
    return bool(CONTAINER_NO_RE.fullmatch(token))


def _is_freight_marker_row(row: list) -> bool:
    """
    Marker row rule:
    - row may or may not contain "CBM FREIGHT/FEIGHT";
    - container number may appear in non-leading columns;
    - if any cell contains a container number like "JXLU4327901",
      treat it as a shipment marker row.
    """
    cells = [(idx, _to_text(v)) for idx, v in enumerate(row) if _to_text(v)]
    if not cells:
        return False
    return any(_is_container_no_token(text) for _, text in cells)


def _header_like_mapping(row: list) -> dict[int, str]:
    keys = {k.replace(" ", "").upper(): v for k, v in HEADER_MAP.items()}
    mapping: dict[int, str] = {}
    for idx, val in enumerate(row):
        kk = _norm_header(val)
        if kk in keys:
            mapping[idx] = keys[kk]
    return mapping


def _extract_batch_date_from_header_row(row: list, mapping: dict[int, str]) -> str | None:
    shop_cols = [idx for idx, field in mapping.items() if field == "shop_no"]
    if not shop_cols:
        return None
    shop_idx = min(shop_cols)
    left_idx = shop_idx - 1
    if left_idx < 0 or left_idx >= len(row):
        return None
    return _to_ymd(row[left_idx])


def _fill_dimensions_from_cbm_suffix(row: list, mapping: dict[int, str], item: dict) -> None:
    """
    Implicit dimension rule:
    length/width/height are read from the 3 cells right after CBM.
    """
    cbm_col = next((idx for idx, field in mapping.items() if field == "cbm_calculated"), None)
    if cbm_col is None:
        return
    if cbm_col + 3 >= len(row):
        return
    item["length_cm"] = to_float(row[cbm_col + 1], 0.0)
    item["width_cm"] = to_float(row[cbm_col + 2], 0.0)
    item["height_cm"] = to_float(row[cbm_col + 3], 0.0)


def _extract_phone_from_book(book: list[tuple[str, list[list]]]) -> str | None:
    c = Counter()
    # Phone is more likely to reflect current customer info in recent sheets.
    for _sname, rows in book[-3:]:
        for row in rows[:1200]:
            if not row:
                continue
            token = _first_col_phone_token(row[0] if len(row) > 0 else None)
            if token and PHONE_RE.fullmatch(token):
                c[token] += 1
    if not c:
        return None
    return c.most_common(1)[0][0]


def _choose_customer_file(folder: Path, min_file_year: int = 0) -> Path | None:
    files = [
        p
        for p in folder.rglob("*")
        if p.is_file()
        and p.suffix.lower() in (".xls", ".xlsx")
        and not p.name.startswith("~$")
    ]
    if min_file_year > 0:
        files = [p for p in files if datetime.fromtimestamp(p.stat().st_mtime).year >= min_file_year]
    if not files:
        return None
    preferred = [p for p in files if "收货清单" in p.name]
    target = preferred or files
    target.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return target[0]


def _collect_customer_folders(data_root: Path) -> list[Path]:
    chinese_start_re = re.compile(r"^[\u4e00-\u9fff]")
    out = []
    for p in sorted(data_root.iterdir()):
        if not p.is_dir():
            continue
        if p.name in EXCLUDED_TOP_DIRS:
            continue
        if chinese_start_re.match(p.name):
            continue
        out.append(p)
    return out


def _parse_in_stock_from_last_sheet(path: Path, customer_name: str) -> tuple[list[ParsedItem], str | None, str, bool]:
    book = _read_book(path, max_cols=40)
    phone = _extract_phone_from_book(book)
    picked = _pick_last_non_empty_sheet(book)
    if not picked:
        return [], phone, "", True
    sheet_name, rows = picked
    header_idx, mapping = _detect_header(rows)
    marker_idx = _last_freight_marker_idx(rows)
    current_batch_date = _extract_batch_date_from_header_row(rows[header_idx], mapping)

    parsed: list[ParsedItem] = []
    last_shop: str | None = None
    last_pos: str | None = None

    for row_no, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        abs_idx = row_no - 1
        if abs_idx <= marker_idx:
            continue

        hm = _header_like_mapping(row)
        if "shop_no" in hm.values() and (("item_no" in hm.values()) or ("item_name_cn" in hm.values())):
            next_date = _extract_batch_date_from_header_row(row, hm)
            if next_date:
                current_batch_date = next_date
            continue

        if _is_freight_marker_row(row):
            continue

        item = {}
        for col_idx, field in mapping.items():
            item[field] = row[col_idx] if col_idx < len(row) else None
        _fill_dimensions_from_cbm_suffix(row, mapping, item)

        item_name = _to_text(item.get("item_name_cn")) or _to_text(item.get("item_no"))
        if not item_name or _is_skip_item_token(item_name):
            continue

        shop_no = _to_text(item.get("shop_no")) or None
        pos = _to_text(item.get("position_or_tel")) or None
        if shop_no:
            last_shop = shop_no
        elif last_shop:
            shop_no = last_shop
        if pos:
            last_pos = pos
        elif last_pos:
            pos = last_pos

        parsed.append(
            ParsedItem(
                customer_name=customer_name,
                shop_no=shop_no,
                position_or_tel=pos,
                item_no=_to_text(item.get("item_no")) or None,
                item_name_cn=item_name,
                material=_to_text(item.get("material")) or None,
                carton_count=to_int(item.get("carton_count"), 0),
                qty=to_int(item.get("qty"), 0),
                unit_price=to_float(item.get("unit_price"), 0.0),
                total_price=to_float(item.get("total_price"), 0.0),
                deposit_hint=to_float(item.get("deposit_hint"), 0.0),
                cbm_calculated=to_float(item.get("cbm_calculated"), 0.0),
                length_cm=to_float(item.get("length_cm"), 0.0),
                width_cm=to_float(item.get("width_cm"), 0.0),
                height_cm=to_float(item.get("height_cm"), 0.0),
                inbound_date_hint=current_batch_date,
                source_file=str(path),
                source_sheet=sheet_name,
                source_row_no=row_no,
            )
        )

    return parsed, phone, sheet_name, False


def _build_inbound_no(source_key: str) -> str:
    h = hashlib.sha1(source_key.encode("utf-8")).hexdigest()[:20].upper()
    return f"HIST-{h}"


def run_import(args) -> dict:
    data_root = Path(args.data_root).resolve()
    if not data_root.exists():
        raise SystemExit(f"data root not found: {data_root}")

    folders = _collect_customer_folders(data_root)
    if args.customer:
        allow = {x.strip() for x in args.customer if x.strip()}
        folders = [f for f in folders if f.name in allow]
    if args.limit and args.limit > 0:
        folders = folders[: args.limit]

    stats = Counter()
    detail_samples = []
    now_date = datetime.now().strftime("%Y-%m-%d")

    with db_session() as conn:
        warehouse = conn.execute("SELECT id FROM warehouses ORDER BY id LIMIT 1").fetchone()
        if not warehouse:
            raise ValueError("warehouse not found")
        warehouse_id = int(warehouse["id"])

        for folder in folders:
            stats["folders_scanned"] += 1
            picked = _choose_customer_file(folder, min_file_year=args.min_file_year)
            if not picked:
                stats["folders_no_file"] += 1
                continue

            # Always ensure customer master exists once a customer file is found,
            # even if this customer has no in-stock rows to import.
            if args.dry_run:
                cid = resolve_customer_id(conn, folder.name)
                if cid is None:
                    stats["customers_would_auto_create"] += 1
            else:
                cid, created = get_or_create_customer_id(conn, folder.name)
                if created:
                    stats["customers_auto_created"] += 1

            try:
                rows, phone, sheet_name, all_sheets_empty = _parse_in_stock_from_last_sheet(picked, folder.name)
            except Exception as e:
                stats["folders_parse_failed"] += 1
                if args.verbose:
                    print(f"[PARSE_FAIL] {folder.name}: {e}")
                continue

            alias_candidates: list[str] = []
            try:
                alias_book = _read_book(picked, max_cols=8)
                alias_candidates, alias_stats = _collect_recent_sheet_aliases(alias_book, folder.name, years=2)
                for k, v in alias_stats.items():
                    stats[k] += v
            except Exception:
                stats["alias_scan_failed"] += 1

            if args.dry_run:
                cid = resolve_customer_id(conn, folder.name)
                if cid is None:
                    # Should not happen because we counted above; skip DB-touch logic in dry-run.
                    if alias_candidates:
                        stats["aliases_detected_unbound_customer"] += len(alias_candidates)
                    if not rows:
                        if all_sheets_empty:
                            stats["folders_all_sheets_empty"] += 1
                            if args.verbose:
                                print(f"[SKIP_ALL_EMPTY_SHEETS] {folder.name}")
                        stats["folders_no_in_stock_rows"] += 1
                    continue

            if alias_candidates:
                if args.dry_run:
                    stats["aliases_would_upsert"] += len(alias_candidates)
                else:
                    for alias_name in alias_candidates:
                        upsert_alias(
                            conn,
                            customer_id=cid,
                            alias_name=alias_name,
                            source="AUTO_DETECT",
                            is_primary=0,
                            is_active=1,
                            remark="auto detected from recent two-year sheets first column",
                        )
                        stats["aliases_upserted"] += 1

            if phone:
                c = conn.execute("SELECT phone FROM customers WHERE id=?", (cid,)).fetchone()
                cur_phone = (c["phone"] or "").strip() if c else ""
                if not cur_phone and not args.dry_run:
                    conn.execute("UPDATE customers SET phone=?, updated_at=datetime('now','localtime') WHERE id=?", (phone, cid))
                    stats["customer_phone_filled"] += 1

            if not rows:
                if all_sheets_empty:
                    stats["folders_all_sheets_empty"] += 1
                    if args.verbose:
                        print(f"[SKIP_ALL_EMPTY_SHEETS] {folder.name}")
                stats["folders_no_in_stock_rows"] += 1
                continue

            file_mtime_date = datetime.fromtimestamp(picked.stat().st_mtime).strftime("%Y-%m-%d")
            for r in rows:
                stats["rows_candidate"] += 1
                source_rel = Path(r.source_file).resolve().relative_to(ROOT.resolve())
                source_key = f"{source_rel}|{r.source_sheet}|R{r.source_row_no}"
                inbound_no = _build_inbound_no(source_key)
                hit = conn.execute("SELECT 1 FROM inbound_items WHERE inbound_no=? LIMIT 1", (inbound_no,)).fetchone()
                if hit:
                    stats["rows_exists_skip"] += 1
                    continue

                payload = {
                    "inbound_no": inbound_no,
                    "import_batch_id": None,
                    "customer_id": cid,
                    "warehouse_id": warehouse_id,
                    "inbound_date": args.inbound_date or r.inbound_date_hint or file_mtime_date,
                    "shop_no": r.shop_no,
                    "position_or_tel": r.position_or_tel,
                    "item_no": r.item_no,
                    "item_name_cn": r.item_name_cn,
                    "material": r.material,
                    "carton_count": r.carton_count,
                    "qty": r.qty,
                    "unit_price": r.unit_price,
                    "total_price": r.total_price,
                    "deposit_hint": r.deposit_hint,
                    "cbm_calculated": r.cbm_calculated,
                    "length_cm": r.length_cm,
                    "width_cm": r.width_cm,
                    "height_cm": r.height_cm,
                    "status": "IN_STOCK",
                    "remark": f"HIST_IMPORT::{source_key}",
                }

                if args.dry_run:
                    stats["rows_would_insert"] += 1
                else:
                    create_inbound_item(conn, payload)
                    stats["rows_inserted"] += 1
                    if len(detail_samples) < 20:
                        detail_samples.append(
                            {
                                "customer": folder.name,
                                "inbound_no": inbound_no,
                                "sheet": sheet_name,
                                "row_no": r.source_row_no,
                                "item_name": r.item_name_cn,
                                "cbm": r.cbm_calculated,
                            }
                        )

            stats["folders_with_rows"] += 1

    mode = "DRY-RUN" if args.dry_run else "APPLY"
    return {
        "mode": mode,
        "data_root": str(data_root),
        "stats": {k: int(v) for k, v in sorted(stats.items(), key=lambda x: x[0])},
        "samples": detail_samples[:10],
    }


def run(args) -> int:
    report = run_import(args)
    print(f"[{report['mode']}] data_root={report['data_root']}")
    for k, v in report["stats"].items():
        print(f"{k}: {v}")
    if report["samples"]:
        print("samples:")
        for x in report["samples"]:
            print(x)
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Import historical in-stock rows from customer files (last sheet only).")
    p.add_argument("--data-root", type=Path, default=DEFAULT_DATA_ROOT, help="Root folder that contains customer folders")
    p.add_argument("--customer", action="append", help="Only import specified customer folder name (repeatable)")
    p.add_argument("--limit", type=int, default=0, help="Limit scanned customer folders")
    p.add_argument("--inbound-date", type=str, default="", help="Override inbound date (YYYY-MM-DD)")
    p.add_argument("--min-file-year", type=int, default=0, help="Only use files with modified year >= this value")
    p.add_argument("--dry-run", action="store_true", default=True, help="Scan and parse without DB insert (default true)")
    p.add_argument("--apply", action="store_true", help="Actually insert rows into DB")
    p.add_argument("--verbose", action="store_true")
    return p


if __name__ == "__main__":
    parser = build_arg_parser()
    args = parser.parse_args()
    if args.apply:
        args.dry_run = False
    raise SystemExit(run(args))
