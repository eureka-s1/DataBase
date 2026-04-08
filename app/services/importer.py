from __future__ import annotations

import json
from datetime import datetime
import re
from pathlib import Path
from sqlite3 import Connection
from uuid import uuid4

import openpyxl
import xlrd

from .common import now_ts, to_float, to_int
from .customers import get_or_create_customer_id, resolve_customer_id
from .inbound import create_inbound_item

HEADER_MAP = {
    'CUSTOMER NAME': 'customer_name',
    'SHOP NO': 'shop_no',
    '位置': 'position_or_tel',
    'TEL': 'position_or_tel',
    'ITEM NO': 'item_no',
    'DESCRIPTION': 'item_name_cn',
    'DESCRPETION': 'item_name_cn',
    'ITEM NAME': 'item_name_cn',
    'ITEMNAME': 'item_name_cn',
    '品名': 'item_name_cn',
    '材质': 'material',
    'CTNS': 'carton_count',
    'CTN': 'carton_count',
    'CNS': 'carton_count',
    'QTY': 'qty',
    'PRICE': 'unit_price',
    'T.PRICE': 'total_price',
    'T.T': 'total_price',
    'AMOUNT': 'total_price',
    '定金': 'deposit_hint',
    'DEP': 'deposit_hint',
    'DEPOZIT': 'deposit_hint',
    'CBM': 'cbm_calculated',
    '长': 'length_cm',
    '宽': 'width_cm',
    '高': 'height_cm',
}


def _to_float_cell(value, default: float = 0.0) -> float:
    s = str(value or '').strip().replace(',', '')
    if not s:
        return default
    try:
        return float(s)
    except Exception:
        m = re.search(r'-?\d+(?:\.\d+)?', s)
        if not m:
            return default
        try:
            return float(m.group(0))
        except Exception:
            return default


def _ensure_receipt_file(path: Path) -> None:
    if '收货清单' not in path.name:
        raise ValueError('仅支持导入文件名包含“收货清单”的入库文件')


def _default_customer_from_path(path: Path) -> str:
    parts = path.parts
    for i, p in enumerate(parts):
        if p == '2026data' and i + 1 < len(parts):
            return str(parts[i + 1]).strip()
    return path.parent.name.strip()


def _is_skip_customer_token(value: str) -> bool:
    v = (value or '').strip().upper()
    if not v:
        return True
    bad = ('TOTAL', 'SEND TO ME', 'BALANCE', 'REMAIN', 'PAID', 'COMMISSION', 'FREIGHT')
    return any(x in v for x in bad)


def _is_skip_item_token(value: str) -> bool:
    v = (value or '').strip().upper()
    if not v:
        return True
    bad_exact = {'品名', 'ITEM NAME', 'ITEM NO', 'SHOP NO', 'TEL', 'TOTAL', 'BALANCE', 'REMAIN'}
    if v in bad_exact:
        return True
    bad_contains = ('CBM FREIGHT', 'SEND TO ME', 'PAID', 'COMMISSION')
    if any(x in v for x in bad_contains):
        return True
    return bool(re.match(r'^\d+\s*CTNS?$', v))


def _read_sheet(path: Path, max_cols: int = 40) -> tuple[str, list[list], list[tuple[int, int, int, int]]]:
    rows: list[list] = []
    merged_ranges: list[tuple[int, int, int, int]] = []
    suffix = path.suffix.lower()
    if suffix == '.xlsx':
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        for rg in ws.merged_cells.ranges:
            merged_ranges.append((int(rg.min_row), int(rg.max_row), int(rg.min_col), int(rg.max_col)))
        for r in ws.iter_rows(values_only=True, max_col=max_cols):
            rows.append(list(r))
        wb.close()
        return sheet_name, rows, merged_ranges
    if suffix == '.xls':
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        for rlo, rhi, clo, chi in sh.merged_cells:
            # xlrd uses 0-based half-open bounds: [rlo, rhi), [clo, chi)
            merged_ranges.append((int(rlo + 1), int(rhi), int(clo + 1), int(chi)))
        for i in range(sh.nrows):
            rows.append(sh.row_values(i, 0, max_cols))
        return sh.name or 'Sheet1', rows, merged_ranges
    raise ValueError('unsupported file type')


def _collapse_rows_by_merged_carton(
    parsed_rows: list[dict],
    merged_ranges: list[tuple[int, int, int, int]],
    carton_col_idx: int | None,
) -> list[dict]:
    if not parsed_rows or carton_col_idx is None:
        return parsed_rows

    carton_col_1based = int(carton_col_idx) + 1
    candidate_ranges = [
        rg for rg in merged_ranges
        if rg[0] < rg[1] and rg[2] <= carton_col_1based <= rg[3]
    ]
    if not candidate_ranges:
        return parsed_rows

    row_to_range: dict[int, tuple[int, int, int, int]] = {}
    for rg in candidate_ranges:
        min_row, max_row, _, _ = rg
        for rn in range(min_row, max_row + 1):
            row_to_range.setdefault(rn, rg)

    rows_by_no = {int(r['row_no']): r for r in parsed_rows}
    sorted_rows = sorted(parsed_rows, key=lambda x: int(x.get('row_no') or 0))
    consumed: set[int] = set()
    out: list[dict] = []

    for row in sorted_rows:
        row_no = int(row.get('row_no') or 0)
        if row_no in consumed:
            continue
        rg = row_to_range.get(row_no)
        if not rg:
            consumed.add(row_no)
            out.append(row)
            continue

        min_row, max_row, _, _ = rg
        members = [rows_by_no[rn] for rn in range(min_row, max_row + 1) if rn in rows_by_no and rn not in consumed]
        if len(members) <= 1:
            consumed.add(row_no)
            out.append(row)
            continue

        base = dict(members[0])
        item_names: list[str] = []
        for m in members:
            name = str(m.get('item_name_cn') or '').strip()
            if name and name not in item_names:
                item_names.append(name)
        if item_names:
            merged_name = '+'.join(item_names)
            if not merged_name.endswith('*'):
                merged_name = f'{merged_name}*'
            base['item_name_cn'] = merged_name

        # Merged-carton rows represent one carton containing multiple item lines:
        # monetary fields should aggregate across member lines.
        base['qty'] = sum(to_int(m.get('qty')) for m in members)
        base['unit_price'] = round(sum(to_float(m.get('unit_price')) for m in members), 2)
        base['total_price'] = round(sum(to_float(m.get('total_price')) for m in members), 2)
        # If CBM cells are not merged, they must still be aggregated by member lines.
        base['cbm_calculated'] = round(sum(to_float(m.get('cbm_calculated')) for m in members), 6)
        base['is_merged_carton'] = True

        raw_row = dict(base.get('raw_row') or {})
        raw_row['MERGED_CARTON_ROWS'] = ','.join(str(int(m.get('row_no') or 0)) for m in members)
        raw_row['MERGED_CARTON'] = 1
        base['raw_row'] = raw_row

        out.append(base)
        consumed.update(int(m.get('row_no') or 0) for m in members)

    return out


def _norm_header(value) -> str:
    return str(value or '').strip().upper().replace(' ', '')


def _build_header_labels(row: list, max_cols: int) -> list[str]:
    labels: list[str] = []
    used: dict[str, int] = {}
    for i in range(max_cols):
        raw = row[i] if i < len(row) else None
        base = str(raw).strip() if raw is not None else ''
        if not base:
            base = f'COL_{i + 1}'
        cnt = used.get(base, 0) + 1
        used[base] = cnt
        labels.append(base if cnt == 1 else f'{base}_{cnt}')
    return labels


def _detect_header(rows: list[list], max_scan: int = 120) -> tuple[int, dict[int, str], list[str]]:
    keys = {k.replace(' ', '').upper(): v for k, v in HEADER_MAP.items()}
    for i, row in enumerate(rows[:max_scan]):
        mapping: dict[int, str] = {}
        for idx, val in enumerate(row):
            kk = _norm_header(val)
            if kk in keys:
                mapping[idx] = keys[kk]
        has_item = ('item_name_cn' in mapping.values()) or ('item_no' in mapping.values())
        if len(mapping) >= 3 and has_item:
            labels = _build_header_labels(row, max(len(row), 20))
            return i, mapping, labels
    raise ValueError('unable to detect inbound header row')


def _row_to_raw_map(row: list, labels: list[str]) -> dict:
    raw: dict = {}
    for i, label in enumerate(labels):
        val = row[i] if i < len(row) else None
        if val is None:
            continue
        s = str(val).strip()
        if s == '':
            continue
        raw[label] = val
    return raw


def _fill_dimensions_from_cbm_suffix(row: list, mapping: dict[int, str], item: dict) -> None:
    """
    Implicit dimension rule:
    if length/width/height are missing, default to the 3 columns right after CBM column.
    """
    cbm_col = next((idx for idx, field in mapping.items() if field == 'cbm_calculated'), None)
    if cbm_col is None:
        return
    if cbm_col + 3 >= len(row):
        return
    # Priority rule:
    # dimensions are primarily read from the three cells right after CBM.
    # if those cells are empty/invalid, fallback to already parsed explicit fields.
    l2 = _to_float_cell(row[cbm_col + 1])
    w2 = _to_float_cell(row[cbm_col + 2])
    h2 = _to_float_cell(row[cbm_col + 3])

    if l2 > 0:
        item['length_cm'] = l2
    elif to_float(item.get('length_cm')) <= 0:
        item['length_cm'] = to_float(item.get('length_cm'))

    if w2 > 0:
        item['width_cm'] = w2
    elif to_float(item.get('width_cm')) <= 0:
        item['width_cm'] = to_float(item.get('width_cm'))

    if h2 > 0:
        item['height_cm'] = h2
    elif to_float(item.get('height_cm')) <= 0:
        item['height_cm'] = to_float(item.get('height_cm'))


def _parse_headerless_excel(path: Path, sheet_name: str, rows: list[list]) -> dict | None:
    parsed = []
    fallback_customer = _default_customer_from_path(path)
    last_customer_name = fallback_customer
    last_shop_by_customer: dict[str, str] = {}

    for i, row in enumerate(rows[:600], start=1):
        vals = [None if v is None else str(v).strip() for v in row]
        while vals and not vals[0]:
            vals.pop(0)
        tokens = [v for v in vals if v]
        if len(tokens) < 5:
            continue

        raw_map = {f'COL_{j + 1}': t for j, t in enumerate(tokens)}
        item = None

        if len(tokens) >= 8 and tokens[3]:
            cbm_a = _to_float_cell(tokens[7]) if len(tokens) > 7 else 0.0
            cbm_a_alt = _to_float_cell(tokens[6]) if len(tokens) > 6 else 0.0
            use_alt = cbm_a <= 0 and cbm_a_alt > 0
            cbm_used = cbm_a_alt if use_alt else cbm_a
            if cbm_used > 0:
                customer = tokens[0] or last_customer_name or fallback_customer
                if not _is_skip_customer_token(customer):
                    last_customer_name = customer
                else:
                    customer = last_customer_name or fallback_customer
                shop_no = tokens[1] if len(tokens) > 1 else None
                shop_no_norm = str(shop_no or '').strip()
                if shop_no_norm:
                    last_shop_by_customer[customer] = shop_no_norm
                elif customer in last_shop_by_customer:
                    shop_no = last_shop_by_customer[customer]
                item = {
                    'customer_name': customer,
                    'shop_no': shop_no,
                    'item_name_cn': tokens[3],
                    'material': tokens[4] if len(tokens) > 4 else None,
                    'carton_count': to_int(tokens[5]) if len(tokens) > 5 else 0,
                    'qty': to_int(tokens[6]) if (len(tokens) > 6 and not use_alt) else 0,
                    'cbm_calculated': cbm_used,
                    'length_cm': _to_float_cell(tokens[7]) if use_alt and len(tokens) > 7 else (_to_float_cell(tokens[8]) if len(tokens) > 8 else 0.0),
                    'width_cm': _to_float_cell(tokens[8]) if use_alt and len(tokens) > 8 else (_to_float_cell(tokens[9]) if len(tokens) > 9 else 0.0),
                    'height_cm': _to_float_cell(tokens[9]) if use_alt and len(tokens) > 9 else (_to_float_cell(tokens[10]) if len(tokens) > 10 else 0.0),
                    'row_no': i,
                    'source_sheet': sheet_name,
                    'raw_row': raw_map,
                }

        if item is None and len(tokens) >= 5 and tokens[0]:
            cbm_b = _to_float_cell(tokens[4]) if len(tokens) > 4 else 0.0
            if cbm_b > 0:
                item = {
                    'customer_name': last_customer_name or fallback_customer,
                    'item_name_cn': tokens[0],
                    'material': tokens[1] if len(tokens) > 1 else None,
                    'carton_count': to_int(tokens[2]) if len(tokens) > 2 else 0,
                    'qty': to_int(tokens[3]) if len(tokens) > 3 else 0,
                    'cbm_calculated': cbm_b,
                    'length_cm': _to_float_cell(tokens[5]) if len(tokens) > 5 else 0.0,
                    'width_cm': _to_float_cell(tokens[6]) if len(tokens) > 6 else 0.0,
                    'height_cm': _to_float_cell(tokens[7]) if len(tokens) > 7 else 0.0,
                    'row_no': i,
                    'source_sheet': sheet_name,
                    'raw_row': raw_map,
                }

        if item and item.get('item_name_cn') and not _is_skip_item_token(str(item.get('item_name_cn'))):
            parsed.append(item)

    if not parsed:
        return None
    return {'sheet_name': sheet_name, 'header_row': 0, 'field_mapping': {}, 'rows': parsed, 'errors': []}


def parse_inbound_excel(path: Path) -> dict:
    _ensure_receipt_file(path)
    sheet_name, rows, merged_ranges = _read_sheet(path)
    try:
        header_idx, mapping, labels = _detect_header(rows)
    except ValueError:
        fallback = _parse_headerless_excel(path, sheet_name, rows)
        if fallback:
            return fallback
        raise

    fallback_customer = _default_customer_from_path(path)
    has_customer_col = 'customer_name' in mapping.values()
    parsed = []
    last_customer_name = ''
    last_shop_by_customer: dict[str, str] = {}

    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        raw_map = _row_to_raw_map(row, labels)
        if not raw_map:
            continue

        item = {}
        for col_idx, field in mapping.items():
            item[field] = row[col_idx] if col_idx < len(row) else None
        _fill_dimensions_from_cbm_suffix(row, mapping, item)

        customer_name = str(item.get('customer_name') or '').strip() if has_customer_col else ''
        if has_customer_col and customer_name and not _is_skip_customer_token(customer_name):
            last_customer_name = customer_name
        else:
            customer_name = last_customer_name
        if not customer_name:
            customer_name = fallback_customer
        shop_no = item.get('shop_no')
        shop_no_norm = str(shop_no or '').strip()
        if shop_no_norm:
            last_shop_by_customer[customer_name] = shop_no_norm
        elif customer_name in last_shop_by_customer:
            shop_no = last_shop_by_customer[customer_name]

        item_name = str(item.get('item_name_cn') or '').strip()
        if not item_name:
            item_name = str(item.get('item_no') or '').strip()
        if not item_name or _is_skip_item_token(item_name):
            continue

        parsed.append(
            {
                'customer_name': customer_name,
                'shop_no': shop_no,
                'position_or_tel': item.get('position_or_tel'),
                'item_no': item.get('item_no'),
                'item_name_cn': item.get('item_name_cn') or item.get('item_no'),
                'material': item.get('material'),
                'carton_count': to_int(item.get('carton_count')),
                'qty': to_int(item.get('qty')),
                'unit_price': to_float(item.get('unit_price')),
                'total_price': to_float(item.get('total_price')),
                'deposit_hint': to_float(item.get('deposit_hint')),
                'length_cm': to_float(item.get('length_cm')),
                'width_cm': to_float(item.get('width_cm')),
                'height_cm': to_float(item.get('height_cm')),
                'cbm_calculated': to_float(item.get('cbm_calculated')),
                'row_no': i,
                'source_sheet': sheet_name,
                'raw_row': raw_map,
            }
        )

    parsed = _collapse_rows_by_merged_carton(
        parsed_rows=parsed,
        merged_ranges=merged_ranges,
        carton_col_idx=next((idx for idx, field in mapping.items() if field == 'carton_count'), None),
    )

    return {
        'sheet_name': sheet_name,
        'header_row': header_idx + 1,
        'field_mapping': {str(k): v for k, v in mapping.items()},
        'rows': parsed,
        'errors': [],
    }


def infer_inbound_date(path: Path, inbound_date: str | None = None) -> str:
    # New rule: do not parse date from filename/path.
    # If user did not provide a date, fallback to system date.
    if inbound_date and str(inbound_date).strip():
        raw = str(inbound_date).strip()
        normalized = (
            raw.replace('年', '-')
            .replace('月', '-')
            .replace('日', '')
            .replace('.', '-')
            .replace('/', '-')
        )
        normalized = re.sub(r'-{2,}', '-', normalized).strip('-')
        try:
            if re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', normalized):
                y, m, d = [int(x) for x in normalized.split('-')]
                return datetime(y, m, d).strftime('%Y-%m-%d')
        except Exception:
            pass
    return datetime.now().strftime('%Y-%m-%d')


def _insert_import_row(
    conn: Connection,
    batch_id: int,
    row_no: int,
    source_sheet: str,
    source_row: dict,
    normalized: dict | None,
    inbound_item_id: int | None,
    is_valid: int,
    error_reason: str | None,
) -> None:
    conn.execute(
        '''
        INSERT OR REPLACE INTO inbound_import_rows(
          import_batch_id, row_no, inbound_item_id, is_valid, error_reason,
          source_sheet, customer_name_raw, item_name_raw, source_row_json, normalized_row_json, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (
            batch_id,
            row_no,
            inbound_item_id,
            is_valid,
            error_reason,
            source_sheet,
            str((source_row or {}).get('CUSTOMER NAME', '') or (source_row or {}).get('COL_1', '')).strip(),
            str((source_row or {}).get('品名', '') or (source_row or {}).get('ITEM NO', '')).strip(),
            json.dumps(source_row or {}, ensure_ascii=False),
            json.dumps(normalized or {}, ensure_ascii=False) if normalized else None,
            now_ts(),
        ),
    )


def import_inbound_excel(conn: Connection, path: Path, inbound_date: str | None, created_by: int, dry_run: bool = True) -> dict:
    parsed = parse_inbound_excel(path)
    ts = now_ts()
    batch_no = f'IB-{ts.replace("-", "").replace(":", "").replace(" ", "")}-{uuid4().hex[:6].upper()}'
    fallback_customer = _default_customer_from_path(path)
    inbound_date_used = infer_inbound_date(path, inbound_date)

    conn.execute(
        '''
        INSERT INTO import_batches(batch_no, source_file, sheet_name, import_type, total_rows, success_rows, failed_rows, created_by, created_at)
        VALUES (?, ?, ?, 'inbound', ?, 0, 0, ?, ?)
        ''',
        (batch_no, str(path), parsed.get('sheet_name') or 'Sheet1', len(parsed['rows']), created_by, ts),
    )
    batch_id = int(conn.execute('SELECT last_insert_rowid()').fetchone()[0])

    success = 0
    failed = 0
    err_rows = []
    auto_created_customers = 0
    auto_created_names: set[str] = set()

    for idx, row in enumerate(parsed['rows'], start=1):
        cid = resolve_customer_id(conn, row['customer_name'])
        if cid is None and fallback_customer:
            cid = resolve_customer_id(conn, fallback_customer)

        normalized = {
            'customer_name': row.get('customer_name'),
            'shop_no': row.get('shop_no'),
            'position_or_tel': row.get('position_or_tel'),
            'item_no': row.get('item_no'),
            'item_name_cn': row.get('item_name_cn'),
            'material': row.get('material'),
            'carton_count': row.get('carton_count'),
            'qty': row.get('qty'),
            'unit_price': row.get('unit_price'),
            'total_price': row.get('total_price'),
            'deposit_hint': row.get('deposit_hint'),
            'length_cm': row.get('length_cm'),
            'width_cm': row.get('width_cm'),
            'height_cm': row.get('height_cm'),
            'cbm_calculated': row.get('cbm_calculated'),
            'inbound_date': inbound_date_used,
            'is_merged_carton': 1 if row.get('is_merged_carton') else 0,
        }

        if cid is None:
            create_name = (row.get('customer_name') or fallback_customer or '').strip()
            if not create_name:
                failed += 1
                reason = "customer name empty and cannot auto-create"
                err_rows.append({'row_no': row['row_no'], 'reason': reason})
                if not dry_run:
                    _insert_import_row(
                        conn,
                        batch_id=batch_id,
                        row_no=int(row['row_no']),
                        source_sheet=row.get('source_sheet') or '',
                        source_row=row.get('raw_row') or {},
                        normalized=normalized,
                        inbound_item_id=None,
                        is_valid=0,
                        error_reason=reason,
                    )
                continue

            if dry_run:
                auto_created_names.add(create_name)
                # keep a placeholder ID for dry-run pass-through
                cid = -1
            else:
                cid, created = get_or_create_customer_id(conn, create_name)
                if created:
                    auto_created_customers += 1
                    auto_created_names.add(create_name)

        payload = {
            'inbound_no': f'IN-{inbound_date_used.replace("-", "")}-{batch_id:06d}-{idx:05d}',
            'import_batch_id': batch_id,
            'customer_id': 1 if cid == -1 else cid,
            'warehouse_id': 1,
            'inbound_date': inbound_date_used,
            'shop_no': row.get('shop_no'),
            'position_or_tel': row.get('position_or_tel'),
            'item_no': row.get('item_no'),
            'item_name_cn': row.get('item_name_cn'),
            'material': row.get('material'),
            'carton_count': row.get('carton_count'),
            'qty': row.get('qty'),
            'unit_price': row.get('unit_price'),
            'total_price': row.get('total_price'),
            'deposit_hint': row.get('deposit_hint'),
            'length_cm': row.get('length_cm'),
            'width_cm': row.get('width_cm'),
            'height_cm': row.get('height_cm'),
            'cbm_calculated': row.get('cbm_calculated'),
            'status': 'IN_STOCK',
            'remark': 'MERGED_CARTON' if row.get('is_merged_carton') else None,
        }

        inbound_item_id = None
        if not dry_run:
            inbound_item_id = create_inbound_item(conn, payload)
            _insert_import_row(
                conn,
                batch_id=batch_id,
                row_no=int(row['row_no']),
                source_sheet=row.get('source_sheet') or '',
                source_row=row.get('raw_row') or {},
                normalized=normalized,
                inbound_item_id=inbound_item_id,
                is_valid=1,
                error_reason=None,
            )
        success += 1

    conn.execute('UPDATE import_batches SET success_rows=?, failed_rows=? WHERE id=?', (success, failed, batch_id))
    return {
        'batch_id': batch_id,
        'sheet_name': parsed.get('sheet_name'),
        'header_row': parsed.get('header_row'),
        'inbound_date_used': inbound_date_used,
        'total_rows': len(parsed['rows']),
        'success_rows': success,
        'failed_rows': failed,
        'auto_created_customer_count': len(auto_created_names) if dry_run else auto_created_customers,
        'auto_created_customers': sorted(auto_created_names),
        'errors': err_rows,
        'dry_run': dry_run,
    }


def rollback_inbound_import_batch(conn: Connection, batch_id: int) -> dict:
    batch = conn.execute(
        'SELECT id, batch_no, import_type, total_rows, success_rows FROM import_batches WHERE id=?',
        (batch_id,),
    ).fetchone()
    if not batch:
        raise ValueError('import batch not found')
    if batch['import_type'] != 'inbound':
        raise ValueError('only inbound import batch can be rolled back')

    locked = conn.execute(
        "SELECT COUNT(*) AS c FROM inbound_items WHERE import_batch_id=? AND status!='IN_STOCK'",
        (batch_id,),
    ).fetchone()
    if int(locked['c']) > 0:
        raise ValueError('batch contains non-IN_STOCK records, rollback denied')

    deleted_rows = conn.execute(
        'DELETE FROM inbound_import_rows WHERE import_batch_id=?',
        (batch_id,),
    ).rowcount
    deleted_items = conn.execute(
        "DELETE FROM inbound_items WHERE import_batch_id=? AND status='IN_STOCK'",
        (batch_id,),
    ).rowcount

    conn.execute(
        'UPDATE import_batches SET error_report_path=? WHERE id=?',
        (f'ROLLED_BACK {now_ts()}', batch_id),
    )

    return {
        'batch_id': int(batch['id']),
        'batch_no': batch['batch_no'],
        'deleted_items': int(deleted_items),
        'deleted_import_rows': int(deleted_rows),
        'message': 'rolled back',
    }


def list_inbound_import_batches(
    conn: Connection,
    limit: int = 100,
    batch_id: int | None = None,
    inbound_date: str | None = None,
) -> list[dict]:
    where = ["b.import_type='inbound'"]
    args: list = []
    if batch_id is not None:
        where.append('b.id=?')
        args.append(batch_id)
    if inbound_date:
        where.append('EXISTS (SELECT 1 FROM inbound_items i2 WHERE i2.import_batch_id=b.id AND i2.inbound_date=?)')
        args.append(inbound_date)
    where_sql = ' AND '.join(where)
    args.append(int(limit))
    rows = conn.execute(
        f'''
        SELECT b.id AS batch_id,
               b.batch_no,
               b.source_file,
               b.total_rows,
               b.success_rows,
               b.failed_rows,
               b.created_at,
               COALESCE((SELECT MIN(i.inbound_date) FROM inbound_items i WHERE i.import_batch_id=b.id), '') AS inbound_date,
               COALESCE((SELECT COUNT(*) FROM inbound_items i WHERE i.import_batch_id=b.id AND i.status='IN_STOCK'), 0) AS in_stock_items,
               COALESCE((SELECT COUNT(*) FROM inbound_items i WHERE i.import_batch_id=b.id), 0) AS current_item_rows
        FROM import_batches b
        WHERE {where_sql}
        ORDER BY b.id DESC
        LIMIT ?
        ''',
        args,
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['revoked_rows'] = max(0, int(d.get('success_rows') or 0) - int(d.get('current_item_rows') or 0))
        result.append(d)
    return result
