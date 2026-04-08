from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
import shutil
from sqlite3 import Connection
import subprocess

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import xlrd

from .common import now_ts, to_float, to_int


_FONT_SONGTI = Font(name="宋体", size=11, bold=False)
_BORDER_THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_FX_CNY = 7.0


def _num(v) -> float:
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _cbm_1(v) -> float:
    return round(_num(v), 1)


def _cny_int(v) -> int:
    return int(round(_num(v) * _FX_CNY))


def _style_songti_center_border(cell) -> None:
    cell.font = _FONT_SONGTI
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER_THIN


def ensure_sync_columns(conn: Connection) -> None:
    cols_ib = {str(r["name"]) for r in conn.execute("PRAGMA table_info(import_batches)").fetchall()}
    if "receipt_synced_at" not in cols_ib:
        conn.execute("ALTER TABLE import_batches ADD COLUMN receipt_synced_at TEXT")
    cols_ct = {str(r["name"]) for r in conn.execute("PRAGMA table_info(containers)").fetchall()}
    if "outbound_synced_at" not in cols_ct:
        conn.execute("ALTER TABLE containers ADD COLUMN outbound_synced_at TEXT")
    if "outbound_customers_synced_at" not in cols_ct:
        conn.execute("ALTER TABLE containers ADD COLUMN outbound_customers_synced_at TEXT")
    if "outbound_manifest_synced_at" not in cols_ct:
        conn.execute("ALTER TABLE containers ADD COLUMN outbound_manifest_synced_at TEXT")


def _find_customer_dir(work_dir: Path, customer_name: str) -> Path | None:
    exact = work_dir / customer_name
    if exact.exists() and exact.is_dir():
        return exact
    up = customer_name.strip().upper()
    for p in work_dir.iterdir():
        if p.is_dir() and p.name.strip().upper() == up:
            return p
    return None


def _normalize_customer_key(name: str) -> str:
    return re.sub(r"\s+", "", str(name or "")).upper().strip()


def _resolve_customer_profile(conn: Connection, customer_id: int | None, fallback_name: str = "") -> tuple[str, list[str]]:
    canonical = str(fallback_name or "").strip()
    aliases: list[str] = []
    cid = int(customer_id or 0)
    if cid > 0:
        row = conn.execute("SELECT name FROM customers WHERE id=? LIMIT 1", (cid,)).fetchone()
        if row and str(row["name"] or "").strip():
            canonical = str(row["name"]).strip()
        ars = conn.execute(
            """
            SELECT alias_name
            FROM customer_aliases
            WHERE customer_id=? AND is_active=1
            ORDER BY is_primary DESC, id ASC
            """,
            (cid,),
        ).fetchall()
        aliases = [str(r["alias_name"]).strip() for r in ars if str(r["alias_name"] or "").strip()]
    elif canonical:
        key = _normalize_customer_key(canonical)
        row = conn.execute(
            """
            SELECT c.id, c.name
            FROM customer_aliases ca
            JOIN customers c ON c.id=ca.customer_id
            WHERE ca.alias_name_norm=? AND ca.is_active=1
            LIMIT 1
            """,
            (key,),
        ).fetchone()
        if not row:
            row = conn.execute(
                "SELECT id, name FROM customers WHERE UPPER(REPLACE(name, ' ', ''))=? LIMIT 1",
                (key,),
            ).fetchone()
        if row:
            cid = int(row["id"])
            canonical = str(row["name"] or canonical).strip()
            ars = conn.execute(
                """
                SELECT alias_name
                FROM customer_aliases
                WHERE customer_id=? AND is_active=1
                ORDER BY is_primary DESC, id ASC
                """,
                (cid,),
            ).fetchall()
            aliases = [str(r["alias_name"]).strip() for r in ars if str(r["alias_name"] or "").strip()]
    seen: set[str] = set()
    candidates: list[str] = []
    for n in [canonical, *aliases, fallback_name]:
        name = str(n or "").strip()
        if not name:
            continue
        k = _normalize_customer_key(name)
        if k in seen:
            continue
        seen.add(k)
        candidates.append(name)
    return canonical, candidates


def _find_customer_dir_from_candidates(work_dir: Path, name_candidates: list[str]) -> Path | None:
    for name in name_candidates:
        exact = work_dir / name
        if exact.exists() and exact.is_dir():
            return exact
    wanted = {_normalize_customer_key(x) for x in name_candidates if str(x or "").strip()}
    if not wanted:
        return None
    for p in work_dir.iterdir():
        if p.is_dir() and _normalize_customer_key(p.name) in wanted:
            return p
    return None


def _find_customer_dirs_from_candidates(work_dir: Path, name_candidates: list[str]) -> list[Path]:
    found: list[Path] = []
    seen: set[str] = set()
    wanted = {_normalize_customer_key(x) for x in name_candidates if str(x or "").strip()}
    if not wanted:
        return []
    for name in name_candidates:
        exact = work_dir / str(name or "").strip()
        if exact.exists() and exact.is_dir():
            rp = str(exact.resolve())
            if rp not in seen:
                seen.add(rp)
                found.append(exact)
    for p in work_dir.iterdir():
        if not p.is_dir():
            continue
        if _normalize_customer_key(p.name) in wanted:
            rp = str(p.resolve())
            if rp not in seen:
                seen.add(rp)
                found.append(p)
    return found


def _precheck_customer_dirs(work_dir: Path, profiles: list[tuple[str, list[str]]]) -> tuple[dict[str, Path], list[str]]:
    """
    Validate all customer->folder mappings before write.
    Returns (resolved_map, errors). Any error means caller should stop the whole sync.
    """
    resolved: dict[str, Path] = {}
    errors: list[str] = []
    for label, candidates in profiles:
        key = _normalize_customer_key(label)
        if key in resolved:
            continue
        matches = _find_customer_dirs_from_candidates(work_dir, candidates)
        if not matches:
            expect = label or (candidates[0] if candidates else "unknown")
            errors.append(f"{expect}: customer folder not found")
            continue
        if len(matches) > 1:
            paths = ", ".join(str(p) for p in matches[:5])
            expect = label or (candidates[0] if candidates else "unknown")
            errors.append(f"{expect}: multiple customer folders matched -> {paths}")
            continue
        resolved[key] = matches[0]
    return resolved, errors


def _pick_receipt_xlsx(customer_dir: Path, customer_name: str) -> Path:
    all_excel = [
        p for p in customer_dir.rglob("*")
        if p.is_file() and p.suffix.lower() in (".xls", ".xlsx") and not p.name.startswith("~$")
    ]
    files = [
        p for p in customer_dir.rglob("*")
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]
    preferred = [p for p in files if "收货清单" in p.name]
    target = preferred or files
    if target:
        target.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return target[0]
    xls_files = [p for p in all_excel if p.suffix.lower() == ".xls"]
    preferred_xls = [p for p in xls_files if "收货清单" in p.name]
    if preferred_xls or xls_files:
        pick_xls = sorted((preferred_xls or xls_files), key=lambda p: p.stat().st_mtime, reverse=True)[0]
        return _convert_xls_to_xlsx(pick_xls, force=True)
    if all_excel:
        names = ", ".join(p.name for p in sorted(all_excel, key=lambda x: x.name)[:3])
        raise ValueError(f"only .xls found (or no writable .xlsx): {names}")
    raise ValueError("no writable .xlsx receipt file found in customer folder")


def _convert_xls_to_xlsx(xls_path: Path, force: bool = False) -> Path:
    out = xls_path.with_suffix(".xlsx")
    if out.exists() and not force:
        return out
    converter = shutil.which("soffice") or shutil.which("libreoffice")
    if converter:
        cmd = [converter, "--headless", "--convert-to", "xlsx", "--outdir", str(xls_path.parent), str(xls_path)]
        res = subprocess.run(cmd, capture_output=True, text=True)
        if res.returncode == 0 and out.exists():
            try:
                wb2 = openpyxl.load_workbook(out)
                if wb2.worksheets:
                    wb2.active = len(wb2.worksheets) - 1
                    wb2.save(out)
            except Exception:
                pass
            return out
    # 无外部转换器时，走纯 Python 样式迁移兜底
    _xls_to_xlsx_preserve_style(xls_path, out)
    return out


def _xls_color_hex(book, idx: int) -> str | None:
    try:
        c = book.colour_map.get(int(idx))
        if not c or len(c) < 3:
            return None
        r, g, b = int(c[0]), int(c[1]), int(c[2])
        return f"{r:02X}{g:02X}{b:02X}"
    except Exception:
        return None


def _xls_border_style(n: int) -> str | None:
    m = {
        0: None,
        1: "thin",
        2: "medium",
        3: "dashed",
        4: "dotted",
        5: "thick",
        6: "double",
        7: "hair",
        8: "mediumDashed",
        9: "dashDot",
        10: "mediumDashDot",
        11: "dashDotDot",
        12: "mediumDashDotDot",
        13: "slantDashDot",
    }
    return m.get(int(n), "thin")


def _xls_to_xlsx_preserve_style(xls_path: Path, out: Path) -> None:
    book = xlrd.open_workbook(str(xls_path), formatting_info=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    h_map = {0: None, 1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify", 6: "centerContinuous", 7: "distributed"}
    v_map = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}

    for sidx in range(book.nsheets):
        sh = book.sheet_by_index(sidx)
        ws = wb.create_sheet(title=(sh.name or f"Sheet{sidx + 1}")[:31])

        # 列宽和行高
        try:
            for c in range(sh.ncols):
                if c in sh.colinfo_map:
                    ci = sh.colinfo_map[c]
                    # xls 宽度单位约为 1/256 字符
                    w = float(ci.width or 0) / 256.0
                    if w > 0:
                        ws.column_dimensions[get_column_letter(c + 1)].width = w
            for r in range(sh.nrows):
                if r in sh.rowinfo_map:
                    ri = sh.rowinfo_map[r]
                    if ri.height and ri.height > 0:
                        # xls 行高单位 twips (1/20 pt)
                        ws.row_dimensions[r + 1].height = float(ri.height) / 20.0
        except Exception:
            pass

        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                xw = ws.cell(r + 1, c + 1)

                # 值
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    pass
                elif cell.ctype == xlrd.XL_CELL_TEXT:
                    xw.value = str(cell.value)
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    v = float(cell.value)
                    iv = int(v)
                    xw.value = iv if v == float(iv) else v
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dtv = xlrd.xldate_as_datetime(cell.value, book.datemode)
                        if (
                            hasattr(dtv, "hour")
                            and hasattr(dtv, "minute")
                            and hasattr(dtv, "second")
                            and dtv.hour == 0
                            and dtv.minute == 0
                            and dtv.second == 0
                        ):
                            xw.value = dtv.date()
                            xw.number_format = "yyyy/mm/dd"
                        else:
                            xw.value = dtv
                    except Exception:
                        xw.value = cell.value
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    xw.value = bool(cell.value)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    xw.value = None
                else:
                    xw.value = cell.value

                # 样式
                try:
                    xf = book.xf_list[sh.cell_xf_index(r, c)]
                except Exception:
                    xf = None
                if not xf:
                    continue

                # 数字格式
                try:
                    fmt = book.format_map.get(xf.format_key)
                    if fmt and getattr(fmt, "format_str", None):
                        xw.number_format = str(fmt.format_str)
                except Exception:
                    pass

                # 字体
                try:
                    f = book.font_list[xf.font_index]
                    color_hex = _xls_color_hex(book, int(getattr(f, "colour_index", 0)))
                    xw.font = Font(
                        name=getattr(f, "name", None) or "Calibri",
                        size=(float(getattr(f, "height", 200)) / 20.0) if getattr(f, "height", None) else 11,
                        bold=bool(getattr(f, "bold", 0)),
                        italic=bool(getattr(f, "italic", 0)),
                        underline="single" if int(getattr(f, "underline_type", 0) or 0) else None,
                        color=color_hex,
                    )
                except Exception:
                    pass

                # 对齐
                try:
                    a = xf.alignment
                    xw.alignment = Alignment(
                        horizontal=h_map.get(int(getattr(a, "hor_align", 0) or 0)),
                        vertical=v_map.get(int(getattr(a, "vert_align", 2) or 2)),
                        wrap_text=bool(getattr(a, "text_wrapped", 0)),
                    )
                except Exception:
                    pass

                # 填充
                try:
                    bg = xf.background
                    pidx = int(getattr(bg, "pattern_colour_index", 0) or 0)
                    fill_hex = _xls_color_hex(book, pidx)
                    if fill_hex:
                        xw.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
                except Exception:
                    pass

                # 边框
                try:
                    b = xf.border
                    left = Side(style=_xls_border_style(int(getattr(b, "left_line_style", 0) or 0)),
                                color=_xls_color_hex(book, int(getattr(b, "left_colour_index", 0) or 0)))
                    right = Side(style=_xls_border_style(int(getattr(b, "right_line_style", 0) or 0)),
                                 color=_xls_color_hex(book, int(getattr(b, "right_colour_index", 0) or 0)))
                    top = Side(style=_xls_border_style(int(getattr(b, "top_line_style", 0) or 0)),
                               color=_xls_color_hex(book, int(getattr(b, "top_colour_index", 0) or 0)))
                    bottom = Side(style=_xls_border_style(int(getattr(b, "bottom_line_style", 0) or 0)),
                                  color=_xls_color_hex(book, int(getattr(b, "bottom_colour_index", 0) or 0)))
                    xw.border = Border(left=left, right=right, top=top, bottom=bottom)
                except Exception:
                    pass

        # 合并单元格
        try:
            for (rlo, rhi, clo, chi) in sh.merged_cells:
                if rhi > rlo and chi > clo:
                    ws.merge_cells(start_row=rlo + 1, end_row=rhi, start_column=clo + 1, end_column=chi)
        except Exception:
            pass

    if wb.worksheets:
        wb.active = len(wb.worksheets) - 1
    wb.save(out)


def _sheet_month_key(sheet_name: str) -> tuple[int, int] | None:
    s = str(sheet_name or "").strip()
    if not s:
        return None
    m = re.search(r"(?<!\d)(20\d{2})\D{0,4}(0?[1-9]|1[0-2])(?:\D|$)", s)
    if not m:
        return None
    try:
        y = int(m.group(1))
        mo = int(m.group(2))
    except Exception:
        return None
    if 1 <= mo <= 12:
        return y, mo
    return None


def _find_month_sheet_name(wb, year: int, month: int) -> str | None:
    target = (int(year), int(month))
    for sname in wb.sheetnames:
        if _sheet_month_key(str(sname)) == target:
            return str(sname)
    return None


def _ensure_month_sheet(wb, customer_name: str, customer_phone: str | None, year: int, month: int):
    existed = _find_month_sheet_name(wb, year, month)
    if existed:
        ws = wb[existed]
        if _sheet_is_blank(ws):
            _write_section_header(ws, 1, customer_name, customer_phone, year, month)
        return ws
    sname = f"{int(year)} {int(month)}"
    ws = wb.create_sheet(title=sname)
    _write_section_header(ws, 1, customer_name, customer_phone, year, month)
    return ws


def _sheet_is_blank(ws) -> bool:
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    if max_row <= 1 and max_col <= 1:
        v = ws.cell(1, 1).value
        return str(v or "").strip() == ""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if str(ws.cell(r, c).value or "").strip():
                return False
    return True


def _write_section_header(ws, row_no: int, customer_name: str, customer_phone: str | None, year: int, month: int) -> None:
    phone_cell = ws.cell(row_no, 1, str(customer_phone or "").strip())
    _style_songti_center_border(phone_cell)
    title = ws.cell(row_no, 2, f"{customer_name} ORDER {year}-{month}")
    title.font = Font(size=24, bold=False)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=11)
    for c in range(2, 12):
        ws.cell(row_no, c).border = _BORDER_THIN
    _style_songti_center_border(ws.cell(row_no, 12, "AUTO"))
    for c in (13, 14, 15):
        _style_songti_center_border(ws.cell(row_no, c, ""))
    headers = ["日期", "SHOP NO", "TEL", "ITEM NO", "品名", "材质", "CTNS", "QTY", "PRICE", "T.PRICE", "定金", "CBM"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row_no + 1, i, h)
        _style_songti_center_border(c)
    for i in (13, 14, 15):
        c = ws.cell(row_no + 1, i, "")
        _style_songti_center_border(c)
    _ensure_column_widths(ws)


def _ensure_column_widths(ws) -> None:
    # Baseline from: 2026data/ONWA/ONWA 2026收货清单.xls (all sheets, median explicit width)
    # Keep fixed widths for stable layout across synced rows.
    width_map = {
        "A": 11.60,
        "B": 9.00,
        "C": 9.40,
        "D": 16.00,
        "E": 9.00,
        "F": 11.60,
        "G": 9.00,
        "H": 9.00,
        "I": 17.30,
        "J": 12.70,
        "K": 12.70,
        "L": 9.40,
        "M": 9.00,
        "N": 9.00,
        "O": 9.00,
    }
    for col, w in width_map.items():
        ws.column_dimensions[col].width = float(w)


_CONTAINER_TOKEN = re.compile(r"(?i)\b(?:[A-Z]{4,5}\d{6,8}|[A-Z]{1,5}\d{5,})\b")


def _row_values(ws, row_no: int, max_col: int = 14) -> list[str]:
    out: list[str] = []
    for c in range(1, max_col + 1):
        v = ws.cell(row_no, c).value
        out.append(str(v or "").strip())
    return out


def _sheet_has_any_content(ws, max_scan_rows: int = 400, max_scan_cols: int = 20) -> bool:
    mr = int(ws.max_row or 0)
    if mr <= 0:
        return False
    for r in range(1, min(mr, max_scan_rows) + 1):
        for c in range(1, max_scan_cols + 1):
            v = ws.cell(r, c).value
            if str(v or "").strip():
                return True
    return False


def _pick_last_non_empty_sheet(wb):
    for s in reversed(wb.sheetnames):
        ws = wb[s]
        if _sheet_has_any_content(ws):
            return ws
    return wb[wb.sheetnames[-1]]


def _is_settlement_marker_values(values: list[str]) -> bool:
    text = " ".join(values)
    return bool(_CONTAINER_TOKEN.search(text))


def _is_header_values(values: list[str]) -> bool:
    joined = " ".join(values).upper()
    return ("SHOP NO" in joined and "ITEM NO" in joined) or ("日期" in joined and "SHOP NO" in joined)


def _apply_customer_settlement_widths_d_to_k(ws) -> None:
    # From customer_receipts_2026data_format_sample.md baseline (D-K)
    width_map = {
        "D": 16.00,
        "E": 9.00,
        "F": 11.60,
        "G": 9.00,
        "H": 9.00,
        "I": 17.30,
        "J": 12.70,
        "K": 12.70,
    }
    for col, w in width_map.items():
        ws.column_dimensions[col].width = float(w)


def _sheet_needs_new_section(ws) -> bool:
    max_row = int(ws.max_row or 0)
    if max_row <= 0:
        return False
    marker_rows: list[int] = []
    for r in range(1, max_row + 1):
        vals = _row_values(ws, r)
        if _is_settlement_marker_values(vals):
            marker_rows.append(r)
    if not marker_rows:
        return False
    last_marker = marker_rows[-1]
    for r in range(last_marker + 1, max_row + 1):
        vals = _row_values(ws, r)
        if not any(vals):
            continue
        if _is_header_values(vals):
            continue
        if _is_settlement_marker_values(vals):
            continue
        return False
    return True


def list_receipt_sync_batches(conn: Connection, limit: int = 100) -> list[dict]:
    ensure_sync_columns(conn)
    rows = conn.execute(
        """
        SELECT b.id AS batch_id, b.batch_no, b.source_file, b.created_at, b.success_rows, b.failed_rows,
               b.receipt_synced_at,
               COALESCE((SELECT COUNT(*) FROM inbound_items i WHERE i.import_batch_id=b.id), 0) AS item_rows
        FROM import_batches b
        WHERE b.import_type='inbound'
        ORDER BY b.id DESC
        LIMIT ?
        """,
        (int(limit),),
    ).fetchall()
    return [dict(r) for r in rows]


def sync_receipts_by_batch(conn: Connection, batch_id: int, work_dir: Path) -> dict:
    ensure_sync_columns(conn)
    rows = conn.execute(
        """
        SELECT i.*, c.name AS customer_name, COALESCE(c.phone, '') AS customer_phone
        FROM inbound_items i
        JOIN customers c ON c.id = i.customer_id
        WHERE i.import_batch_id=?
        ORDER BY c.name, i.inbound_date, i.id
        """,
        (batch_id,),
    ).fetchall()
    if not rows:
        return {"batch_id": batch_id, "customers": 0, "rows": 0, "files_updated": 0, "errors": ["no inbound rows in batch"]}

    grouped: dict[int, list] = {}
    for r in rows:
        grouped.setdefault(int(r["customer_id"] or 0), []).append(r)

    files_updated = 0
    updated_files: list[dict] = []
    err: list[str] = []
    row_count = 0
    row_ptr: dict[tuple[int, int, int], int] = {}
    first_meta_written: dict[tuple[int, int, int], int] = {}
    grouped_items = sorted(
        grouped.items(),
        key=lambda kv: str(kv[1][0]["customer_name"] or "").strip().upper(),
    )
    profiles: list[tuple[str, list[str]]] = []
    id_to_profile: dict[int, tuple[str, list[str]]] = {}
    for customer_id, items in grouped_items:
        fallback_name = str(items[0]["customer_name"] or "").strip()
        customer_name, name_candidates = _resolve_customer_profile(conn, customer_id, fallback_name=fallback_name)
        label = customer_name or fallback_name or f"customer_id={customer_id}"
        profiles.append((label, name_candidates))
        id_to_profile[customer_id] = (label, name_candidates)
    resolved_dirs, pre_errors = _precheck_customer_dirs(work_dir, profiles)
    if pre_errors:
        return {
            "batch_id": batch_id,
            "customers": len(grouped),
            "rows": 0,
            "files_updated": 0,
            "updated_files": [],
            "errors": pre_errors,
        }

    for customer_id, items in grouped_items:
        fallback_name = str(items[0]["customer_name"] or "").strip()
        customer_name, _name_candidates = id_to_profile.get(customer_id, (fallback_name, [fallback_name]))
        cdir = resolved_dirs.get(_normalize_customer_key(customer_name))
        if not cdir:
            err.append(f"{customer_name}: customer folder precheck failed")
            continue
        try:
            xlsx = _pick_receipt_xlsx(cdir, customer_name)
            wb = openpyxl.load_workbook(xlsx)
            phone = str(items[0]["customer_phone"] or "").strip()
            changed = False
            items_sorted = sorted(
                items,
                key=lambda x: (
                    str(x["shop_no"] or "").strip().upper(),
                    str(x["inbound_date"] or ""),
                    int(x["id"] or 0),
                ),
            )
            last_shop_by_key: dict[tuple[int, int, int], str] = {}
            for it in items_sorted:
                try:
                    d = datetime.strptime(str(it["inbound_date"]), "%Y-%m-%d")
                except Exception:
                    d = datetime.now()
                ws = _ensure_month_sheet(wb, customer_name, phone, d.year, d.month)
                key = (customer_id, d.year, d.month)
                if key not in row_ptr:
                    if _sheet_needs_new_section(ws):
                        start = ws.max_row + 3
                        _write_section_header(ws, start, customer_name, phone, d.year, d.month)
                        row_ptr[key] = start + 2
                    else:
                        row_ptr[key] = ws.max_row + 2
                cur_shop = str(it["shop_no"] or "").strip().upper()
                prev_shop = last_shop_by_key.get(key)
                if prev_shop is not None and cur_shop != prev_shop:
                    row_ptr[key] += 1
                rno = row_ptr[key]
                row_ptr[key] = rno + 1
                last_shop_by_key[key] = cur_shop
                if key not in first_meta_written:
                    col1 = customer_name
                    first_meta_written[key] = 1
                elif first_meta_written[key] == 1:
                    col1 = d.strftime("%Y/%m/%d")
                    first_meta_written[key] = 2
                else:
                    col1 = ""
                vals = [
                    col1,
                    it["shop_no"] or "",
                    it["position_or_tel"] or "",
                    it["item_no"] or "",
                    it["item_name_cn"] or "",
                    it["material"] or "",
                    to_int(it["carton_count"], 0),
                    to_int(it["qty"], 0),
                    to_float(it["unit_price"], 0.0),
                    to_float(it["total_price"], 0.0),
                    to_float(it["deposit_hint"], 0.0),
                    to_float(it["cbm_override"], to_float(it["cbm_calculated"], 0.0)),
                    to_float(it["length_cm"], 0.0),
                    to_float(it["width_cm"], 0.0),
                    to_float(it["height_cm"], 0.0),
                ]
                for cno, v in enumerate(vals, start=1):
                    cell = ws.cell(rno, cno, v)
                    _style_songti_center_border(cell)
                ws.cell(rno, 9).number_format = "$#,##0"
                ws.cell(rno, 10).number_format = "$#,##0"
                ws.cell(rno, 11).number_format = "$#,##0"
                ws.cell(rno, 12).number_format = "0.0"
                row_count += 1
                changed = True
            if changed:
                if wb.worksheets:
                    wb.active = len(wb.worksheets) - 1
                wb.save(xlsx)
                files_updated += 1
                updated_files.append({"customer_name": customer_name, "file_path": str(xlsx)})
        except Exception as e:
            err.append(f"{customer_name}: {e}")

    conn.execute("UPDATE import_batches SET receipt_synced_at=? WHERE id=?", (now_ts(), batch_id))
    return {
        "batch_id": batch_id,
        "customers": len(grouped),
        "rows": row_count,
        "files_updated": files_updated,
        "updated_files": updated_files,
        "errors": err,
    }


def list_outbound_sync_containers(conn: Connection, limit: int = 200) -> list[dict]:
    ensure_sync_columns(conn)
    rows = conn.execute(
        """
        SELECT c.id, c.container_no, c.status, c.confirmed_at, c.outbound_synced_at,
               c.outbound_customers_synced_at, c.outbound_manifest_synced_at,
               COALESCE((SELECT COUNT(*) FROM container_items ci WHERE ci.container_id=c.id), 0) AS item_count
        FROM containers c
        WHERE c.status='CONFIRMED'
        ORDER BY c.id DESC
        LIMIT ?
        """,
        (int(limit),),
    ).fetchall()
    return [dict(r) for r in rows]


def _get_outbound_container_customer_rows(conn: Connection, container_id: int):
    ensure_sync_columns(conn)
    c = conn.execute("SELECT * FROM containers WHERE id=? AND status='CONFIRMED'", (container_id,)).fetchone()
    if not c:
        raise ValueError("container not found or not CONFIRMED")
    rows = conn.execute(
        """
        SELECT cu.id AS customer_id, cu.name AS customer_name, COALESCE(cu.phone, '') AS customer_phone,
               SUM(COALESCE(i.carton_count,0)) AS ctns,
               SUM(COALESCE(ci.cbm_at_load, COALESCE(i.cbm_override, i.cbm_calculated), 0)) AS cbm_total
        FROM container_items ci
        JOIN inbound_items i ON i.id=ci.inbound_item_id
        JOIN customers cu ON cu.id=i.customer_id
        WHERE ci.container_id=?
        GROUP BY cu.id, cu.name, cu.phone
        ORDER BY cu.name
        """,
        (container_id,),
    ).fetchall()
    return c, rows


def sync_outbound_container_to_customers(conn: Connection, container_id: int, work_dir: Path) -> dict:
    c, rows = _get_outbound_container_customer_rows(conn, container_id)
    if not rows:
        return {"container_id": container_id, "customers": 0, "files_updated": 0, "errors": ["container has no items"]}

    files_updated = 0
    updated_files: list[dict] = []
    err: list[str] = []
    date_text = datetime.now().strftime("%Y/%m/%d")
    unit_price = to_float(c["default_price_per_m3"], 0.0)
    fx_cny = _FX_CNY

    profiles: list[tuple[str, list[str]]] = []
    id_to_profile: dict[int, tuple[str, list[str]]] = {}
    for r in rows:
        customer_id = int(r["customer_id"] or 0)
        fallback_name = str(r["customer_name"] or "").strip()
        name, name_candidates = _resolve_customer_profile(conn, customer_id, fallback_name=fallback_name)
        label = name or fallback_name or f"customer_id={customer_id}"
        profiles.append((label, name_candidates))
        id_to_profile[customer_id] = (label, name_candidates)
    resolved_dirs, pre_errors = _precheck_customer_dirs(work_dir, profiles)
    if pre_errors:
        return {
            "container_id": container_id,
            "customers": len(rows),
            "files_updated": 0,
            "updated_files": [],
            "errors": pre_errors,
        }

    for r in rows:
        customer_id = int(r["customer_id"] or 0)
        fallback_name = str(r["customer_name"] or "").strip()
        name, _name_candidates = id_to_profile.get(customer_id, (fallback_name, [fallback_name]))
        cdir = resolved_dirs.get(_normalize_customer_key(name))
        if not cdir:
            err.append(f"{name}: customer folder precheck failed")
            continue
        try:
            xlsx = _pick_receipt_xlsx(cdir, name)
            wb = openpyxl.load_workbook(xlsx)
            ws = _pick_last_non_empty_sheet(wb)
            _apply_customer_settlement_widths_d_to_k(ws)
            row_no = ws.max_row + 3
            ctns = to_int(r["ctns"], 0)
            cbm = _cbm_1(to_float(r["cbm_total"], 0.0))
            freight_usd = round(cbm * unit_price, 2)
            freight_cny = int(round(freight_usd * fx_cny))
            vals = {
                4: str(c["container_no"] or ""),
                5: f"{ctns} CTNS",
                6: str(r["customer_phone"] or ""),
                7: name,
                8: date_text,
                9: f"{cbm:.1f} CBM FREIGHT",
                10: freight_cny,
                11: freight_usd,
            }
            for cno, v in vals.items():
                cell = ws.cell(row_no, cno, v)
                cell.font = Font(name="宋体", size=11, bold=False)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER_THIN
            ws.cell(row_no, 10).number_format = "¥#,##0"
            ws.cell(row_no, 11).number_format = "$#,##0"
            if wb.worksheets:
                wb.active = len(wb.worksheets) - 1
            wb.save(xlsx)
            files_updated += 1
            updated_files.append({"customer_name": name, "file_path": str(xlsx)})
        except Exception as e:
            err.append(f"{name}: {e}")

    ts = now_ts()
    conn.execute(
        "UPDATE containers SET outbound_synced_at=?, outbound_customers_synced_at=?, updated_at=? WHERE id=?",
        (ts, ts, ts, container_id),
    )
    return {
        "container_id": container_id,
        "customers": len(rows),
        "files_updated": files_updated,
        "updated_files": updated_files,
        "errors": err,
    }


def _choose_month_manifest_file(work_dir: Path, year: int, month: int, allow_create: bool) -> tuple[Path | None, bool, str]:
    manifest_dir = work_dir / "装柜清单"
    manifest_dir.mkdir(parents=True, exist_ok=True)
    candidates = [
        manifest_dir / f"{year} {month}月.xlsx",
        manifest_dir / f"{year} {month}.xlsx",
        manifest_dir / f"{year} {month}月.xls",
        manifest_dir / f"{year} {month}.xls",
    ]
    for p in candidates:
        if p.exists() and p.suffix.lower() == ".xlsx":
            return p, False, ""
        if p.exists() and p.suffix.lower() == ".xls":
            x = _convert_xls_to_xlsx(p, force=True)
            return x, False, ""
    create_target = manifest_dir / f"{year} {month}月.xlsx"
    if not allow_create:
        return None, True, str(create_target)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(create_target)
    return create_target, False, ""


def sync_outbound_container_to_manifest(conn: Connection, container_id: int, work_dir: Path, allow_create: bool = False) -> dict:
    c, rows = _get_outbound_container_customer_rows(conn, container_id)
    if not rows:
        return {"container_id": container_id, "customers": 0, "files_updated": 0, "errors": ["container has no items"]}

    try:
        dt = datetime.strptime(str(c["confirmed_at"] or "")[:10], "%Y-%m-%d")
    except Exception:
        dt = datetime.now()
    manifest_file, need_create, suggest = _choose_month_manifest_file(work_dir, dt.year, dt.month, allow_create=allow_create)
    if need_create:
        return {
            "ok": False,
            "need_create": True,
            "message": "monthly manifest file not found",
            "suggest_create_path": suggest,
            "container_id": container_id,
        }
    if not manifest_file:
        raise ValueError("monthly manifest file not found")

    # master customer first
    master_id = c["master_customer_id"]
    if master_id:
        by_name = {str(r["customer_name"]): r for r in rows}
        mrow = conn.execute("SELECT name FROM customers WHERE id=?", (master_id,)).fetchone()
        if mrow and str(mrow["name"]) in by_name:
            master_name = str(mrow["name"])
            rest = [r for r in rows if str(r["customer_name"]) != master_name]
            rest.sort(key=lambda x: str(x["customer_name"] or ""))
            rows_sorted = [by_name[master_name]] + rest
        else:
            rows_sorted = sorted(rows, key=lambda x: str(x["customer_name"] or ""))
    else:
        rows_sorted = sorted(rows, key=lambda x: str(x["customer_name"] or ""))

    wb = openpyxl.load_workbook(manifest_file)
    ws = wb[wb.sheetnames[0]] if wb.sheetnames else wb.create_sheet("Sheet1")
    start = ws.max_row + 3 if ws.max_row > 0 else 1
    date_text = dt.strftime("%Y/%m/%d")
    unit_price = to_float(c["default_price_per_m3"], 0.0)
    container_capacity = to_float(c["capacity_cbm"], 68.0)
    total_freight = round(container_capacity * unit_price, 2)

    # row 1
    head_vals = {1: date_text, 2: str(c["container_no"] or ""), 3: "", 4: "", 5: total_freight}
    for cno, v in head_vals.items():
        cell = ws.cell(start, cno, v)
        cell.font = Font(name="宋体", size=11, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_THIN
    ws.cell(start, 5).number_format = "$#,##0"
    # row 2 header
    hdr = ["PHONE NUMBER", "NAME", "CBM", "CTN", "FREIGHT"]
    for idx, h in enumerate(hdr, start=1):
        cell = ws.cell(start + 1, idx, h)
        cell.font = Font(name="宋体", size=11, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_THIN
    # rows customers
    for i, r in enumerate(rows_sorted, start=0):
        rr = start + 2 + i
        ctns = to_int(r["ctns"], 0)
        cbm = _cbm_1(to_float(r["cbm_total"], 0.0))
        freight = round(cbm * unit_price, 2)
        vals = [str(r["customer_phone"] or ""), str(r["customer_name"] or ""), cbm, ctns, freight]
        for cno, v in enumerate(vals, start=1):
            cell = ws.cell(rr, cno, v)
            cell.font = Font(name="宋体", size=11, bold=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_THIN
        ws.cell(rr, 5).number_format = "$#,##0"

    # keep base widths from 2026 4月.xlsx
    ws.column_dimensions["A"].width = 14.4444
    ws.column_dimensions["B"].width = 19.7778
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 14.3333
    if wb.worksheets:
        wb.active = 0
    wb.save(manifest_file)

    ts = now_ts()
    conn.execute(
        "UPDATE containers SET outbound_manifest_synced_at=?, updated_at=? WHERE id=?",
        (ts, ts, container_id),
    )
    return {
        "ok": True,
        "container_id": container_id,
        "customers": len(rows_sorted),
        "files_updated": 1,
        "updated_files": [{"file_path": str(manifest_file), "customer_name": "装柜清单"}],
        "errors": [],
    }


def sync_outbound_container(conn: Connection, container_id: int, work_dir: Path) -> dict:
    # backward-compatible: keep old endpoint syncing customer files only
    return sync_outbound_container_to_customers(conn, container_id, work_dir)


@dataclass
class MonthlyUpdateResult:
    ym: str
    files_scanned: int
    files_updated: int
    errors: list[str]


def monthly_create_sheet(work_dir: Path, year: int, month: int) -> MonthlyUpdateResult:
    ym = f"{year:04d} {month}"
    files = [
        p for p in work_dir.rglob("*")
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$") and "收货清单" in p.name
    ]
    updated = 0
    errors: list[str] = []
    for p in files:
        try:
            wb = openpyxl.load_workbook(p)
            existed = _find_month_sheet_name(wb, year, month)
            if not existed:
                wb.create_sheet(title=ym)
                if wb.worksheets:
                    wb.active = len(wb.worksheets) - 1
                wb.save(p)
                updated += 1
        except Exception as e:
            errors.append(f"{p}: {e}")
    return MonthlyUpdateResult(ym=ym, files_scanned=len(files), files_updated=updated, errors=errors)
