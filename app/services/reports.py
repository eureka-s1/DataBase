from __future__ import annotations

from pathlib import Path
from sqlite3 import Connection

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

from .common import today_str
from .containers import container_manifest
from .finance import ledger
from .inbound import list_inbound

FX_CNY = 7.0


def _num(v) -> float:
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _cbm_1(v) -> float:
    return round(_num(v), 1)


def _cny_int(v) -> int:
    return int(round(_num(v) * FX_CNY))


def _cny(v) -> float:
    return _num(v) * FX_CNY


def _group_key_text(v) -> str:
    return str(v or '').strip()


def _append_grouped_rows_with_blanks(ws, rows: list[list], customer_idx: int, shop_idx: int) -> None:
    """
    Rows are expected to be pre-sorted by customer -> shop.
    Rule:
    - same SHOP NO block ends -> append 1 blank row
    - same customer block ends -> append 1 extra blank row
    """
    prev_customer: str | None = None
    prev_shop: str | None = None
    for rec in rows:
        cust = _group_key_text(rec[customer_idx] if customer_idx < len(rec) else '')
        shop = _group_key_text(rec[shop_idx] if shop_idx < len(rec) else '')
        if prev_customer is not None:
            if cust != prev_customer:
                ws.append([])
                ws.append([])
            elif shop != prev_shop:
                ws.append([])
        ws.append(rec)
        prev_customer = cust
        prev_shop = shop
    if prev_customer is not None:
        ws.append([])
        ws.append([])


def _receipt_col_widths() -> list[float]:
    # same as receipt-sync A-O widths
    # add one leading "inbound_date" column
    return [11.60, 11.60, 9.00, 9.40, 16.00, 9.00, 11.60, 9.00, 9.00, 17.30, 12.70, 12.70, 9.40, 9.00, 9.00, 9.00]


def _receipt_goods_headers() -> list[str]:
    # align with receipt-sync field order, no "status"
    return ['入库日期', '客户', 'SHOP NO', 'TEL', 'ITEM NO', '品名', '材质', 'CTNS', 'QTY', 'PRICE', 'T.PRICE', '定金', 'CBM', '长', '宽', '高']


def _customer_rows_sync_like(rows: list[dict], master_customer_id: int | None = None) -> list[dict]:
    by_name = {str(r.get('customer_name') or ''): r for r in rows}
    sorted_rows = sorted(rows, key=lambda x: _group_key_text(x.get('customer_name')).upper())
    if not master_customer_id:
        return sorted_rows
    master_row = next((r for r in rows if int(r.get('customer_id') or 0) == int(master_customer_id)), None)
    if not master_row:
        return sorted_rows
    rest = [r for r in sorted_rows if r is not master_row]
    return [master_row] + rest


def _format_receipt_goods_sheet(ws, header_row: int = 1) -> None:
    for r in range(header_row + 1, ws.max_row + 1):
        # skip blank separator rows
        if all(ws.cell(r, c).value in (None, "") for c in range(1, 17)):
            continue
        ws.cell(r, 8).number_format = "0"
        ws.cell(r, 9).number_format = "0"
        ws.cell(r, 10).number_format = "$#,##0"  # display rounded integer USD
        ws.cell(r, 11).number_format = "$#,##0"
        ws.cell(r, 12).number_format = "$#,##0"
        ws.cell(r, 13).number_format = "0.0"      # CBM 1 decimal display
        ws.cell(r, 14).number_format = "0.###"
        ws.cell(r, 15).number_format = "0.###"
        ws.cell(r, 16).number_format = "0.###"


def _format_customer_block_sheet(ws, header_row: int, data_start_row: int) -> None:
    for r in range(data_start_row, ws.max_row + 1):
        if all(ws.cell(r, c).value in (None, "") for c in range(1, 6)):
            continue
        ws.cell(r, 3).number_format = "0.0"
        ws.cell(r, 4).number_format = "0"
        ws.cell(r, 5).number_format = "$#,##0"


def _export_dir() -> Path:
    path = Path('exports')
    path.mkdir(parents=True, exist_ok=True)
    return path


def _sanitize_filename(s: str) -> str:
    bad = '\\/:*?"<>|'
    out = ''.join('_' if ch in bad else ch for ch in (s or ''))
    return out.strip() or 'export'


def _style_sheet(ws, col_widths: list[int] | None = None) -> None:
    if ws.max_row <= 0:
        return
    header_fill = PatternFill('solid', fgColor='E8F1FB')
    thin = Side(style='thin', color='D9E1EA')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='1F2D3D')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical='center')
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.freeze_panes = 'A2'
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col in range(1, ws.max_column + 1):
            max_len = 8
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)


def _new_pdf(out: Path):
    try:
        pdfmetrics.getFont('STSong-Light')
    except Exception:
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    c = canvas.Canvas(str(out), pagesize=A4)
    c.setFont('STSong-Light', 11)
    return c


def export_daily_inbound_excel(conn: Connection, inbound_date: str | None = None) -> str:
    d = inbound_date or today_str()
    data = list_inbound(conn, inbound_date=d)

    wb = Workbook()
    ws = wb.active
    ws.title = 'daily_inbound'
    ws.append(_receipt_goods_headers())
    recs: list[list] = []
    for row in data:
        recs.append([
            row.get('inbound_date'), row.get('customer_name'), row.get('shop_no'), row.get('position_or_tel'),
            row.get('item_no'), row.get('item_name_cn'), row.get('material'), row.get('carton_count'),
            row.get('qty'), _num(row.get('unit_price')), _num(row.get('total_price')),
            _num(row.get('deposit_hint')), _num(row.get('cbm_final')),
            row.get('length_cm'), row.get('width_cm'), row.get('height_cm'),
        ])
    recs.sort(key=lambda r: (_group_key_text(r[1]).upper(), _group_key_text(r[2]).upper(), _group_key_text(r[4]).upper()))
    _append_grouped_rows_with_blanks(ws, recs, customer_idx=1, shop_idx=2)
    _style_sheet(ws, _receipt_col_widths())
    _format_receipt_goods_sheet(ws, header_row=1)

    out = _export_dir() / f'daily_inbound_{d}.xlsx'
    wb.save(out)
    return str(out)


def export_inventory_excel(conn: Connection) -> str:
    data = list_inbound(conn, only_in_stock=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'inventory'
    ws.append(_receipt_goods_headers())
    recs: list[list] = []
    for row in data:
        recs.append([
            row.get('inbound_date'), row.get('customer_name'), row.get('shop_no'), row.get('position_or_tel'),
            row.get('item_no'), row.get('item_name_cn'), row.get('material'), row.get('carton_count'),
            row.get('qty'), _num(row.get('unit_price')), _num(row.get('total_price')),
            _num(row.get('deposit_hint')), _num(row.get('cbm_final')),
            row.get('length_cm'), row.get('width_cm'), row.get('height_cm'),
        ])
    recs.sort(key=lambda r: (_group_key_text(r[1]).upper(), _group_key_text(r[2]).upper(), _group_key_text(r[4]).upper()))
    _append_grouped_rows_with_blanks(ws, recs, customer_idx=1, shop_idx=2)
    _style_sheet(ws, _receipt_col_widths())
    _format_receipt_goods_sheet(ws, header_row=1)
    out = _export_dir() / f'inventory_{today_str()}.xlsx'
    wb.save(out)
    return str(out)


def export_ledger_excel(conn: Connection) -> str:
    data = ledger(conn)
    wb = Workbook()
    ws = wb.active
    ws.title = 'ledger'
    ws.append(['customer_id', 'name', 'aliases', 'total_deposit', 'total_freight', 'total_due', 'net_balance'])
    for row in data:
        ws.append([
            row.get('customer_id'), row.get('name'), ' / '.join(row.get('aliases') or []), _num(row.get('total_deposit')),
            _num(row.get('total_freight')), _num(row.get('total_due')), _num(row.get('net_balance')),
        ])
    for r in range(2, ws.max_row + 1):
        for c in (4, 5, 6, 7):
            ws.cell(r, c).number_format = "$#,##0"
    _style_sheet(ws, [10, 18, 24, 12, 12, 12, 12])
    out = _export_dir() / f'ledger_{today_str()}.xlsx'
    wb.save(out)
    return str(out)


def statement_lines(conn: Connection, statement_id: int) -> tuple[dict, list[dict]]:
    head = conn.execute(
        '''
        SELECT s.id, s.statement_no, s.statement_date, s.status, s.container_id,
               c.container_no, c.master_customer_id
        FROM settlement_statements s
        JOIN containers c ON c.id = s.container_id
        WHERE s.id=?
        ''',
        (statement_id,),
    ).fetchone()
    if not head:
        raise ValueError('statement not found')

    rows = conn.execute(
        '''
        SELECT sl.*, cu.name AS customer_name
        FROM settlement_lines sl
        JOIN customers cu ON cu.id = sl.customer_id
        WHERE sl.statement_id=?
        ORDER BY cu.name
        ''',
        (statement_id,),
    ).fetchall()
    return dict(head), [dict(r) for r in rows]


def export_statement_excel(conn: Connection, statement_id: int) -> str:
    head, rows = statement_lines(conn, statement_id)
    c_head, _items, csum = container_manifest(conn, int(head.get('container_id')))
    csum_sorted = _customer_rows_sync_like(csum, master_customer_id=int(c_head.get('master_customer_id') or 0))
    wb = Workbook()
    ws = wb.active
    ws.title = 'statement'
    ws.append(['日期', '柜号', '结算单号', '状态', '运费'])
    ws.append([
        head.get('statement_date'),
        head.get('container_no'),
        head.get('statement_no'),
        head.get('status'),
        _num(c_head.get('default_price_per_m3')) * _num(c_head.get('capacity_cbm')),
    ])
    ws.cell(2, 5).number_format = "$#,##0"
    ws.append([])
    ws.append(['PHONE NUMBER', 'NAME', 'CBM', 'CTN', 'FREIGHT'])
    customer_header_row = ws.max_row
    for r in csum_sorted:
        ws.append([
            r.get('customer_phone') or '',
            r.get('customer_name') or '',
            _num(r.get('cbm_total')),
            _num(r.get('ctns')),
            _num(r.get('freight_amount')),
        ])
    _format_customer_block_sheet(ws, header_row=customer_header_row, data_start_row=customer_header_row + 1)
    ws.append([])
    ws.append(['NAME', 'DEPOSIT_USED', 'AMOUNT_DUE', 'BALANCE'])
    settle_header_row = ws.max_row
    rows_sorted = sorted(rows, key=lambda x: _group_key_text(x.get('customer_name')).upper())
    for r in rows_sorted:
        ws.append([r.get('customer_name') or '', _num(r.get('deposit_used')), _num(r.get('amount_due')), _num(r.get('amount_balance'))])
    for r in range(settle_header_row + 1, ws.max_row + 1):
        ws.cell(r, 2).number_format = "$#,##0"
        ws.cell(r, 3).number_format = "$#,##0"
        ws.cell(r, 4).number_format = "$#,##0"
    _style_sheet(ws, [14, 20, 12, 10, 12])
    out = _export_dir() / f"statement_{_sanitize_filename(head['statement_no'])}.xlsx"
    wb.save(out)
    return str(out)


def export_statement_pdf(conn: Connection, statement_id: int) -> str:
    head, rows = statement_lines(conn, statement_id)
    c_head, _items, csum = container_manifest(conn, int(head.get('container_id')))
    csum_sorted = _customer_rows_sync_like(csum, master_customer_id=int(c_head.get('master_customer_id') or 0))
    out = _export_dir() / f"statement_{_sanitize_filename(head['statement_no'])}.pdf"
    c = _new_pdf(out)
    width, height = A4
    y = height - 40
    c.drawString(40, y, f"结算单: {head['statement_no']}  柜号: {head['container_no']}  日期: {head['statement_date']}")
    y -= 30
    c.drawString(40, y, 'PHONE')
    c.drawString(150, y, 'NAME')
    c.drawString(300, y, 'CBM')
    c.drawString(360, y, 'CTN')
    c.drawString(430, y, 'FREIGHT(CNY)')
    y -= 16
    for r in csum_sorted:
        if y < 60:
            c.showPage()
            y = height - 40
            c.setFont('STSong-Light', 11)
        c.drawString(40, y, str(r.get('customer_phone') or '')[:12])
        c.drawString(150, y, str(r.get('customer_name') or '')[:18])
        c.drawRightString(345, y, f"{_cbm_1(r.get('cbm_total')):.1f}")
        c.drawRightString(390, y, f"{int(r.get('ctns') or 0)}")
        c.drawRightString(530, y, f"{int(round(_num(r.get('freight_amount'))))}")
        y -= 14
    c.save()
    return str(out)


def export_container_excel(conn: Connection, container_id: int) -> str:
    head, items, customer_summary = container_manifest(conn, container_id)
    csum_sorted = _customer_rows_sync_like(customer_summary, master_customer_id=int(head.get('master_customer_id') or 0))
    wb = Workbook()
    ws = wb.active
    ws.title = 'container_manifest'
    ws.append(['日期', '柜号', '', '', '运费'])
    ws.append([today_str(), head.get('container_no'), '', '', _num(head.get('default_price_per_m3')) * _num(head.get('capacity_cbm'))])
    ws.cell(2, 5).number_format = "$#,##0"
    ws.append([])
    ws.append(['PHONE NUMBER', 'NAME', 'CBM', 'CTN', 'FREIGHT'])
    customer_header_row = ws.max_row
    for r in csum_sorted:
        ws.append([
            r.get('customer_phone') or '',
            r.get('customer_name') or '',
            _num(r.get('cbm_total')),
            _num(r.get('ctns')),
            _num(r.get('freight_amount')),
        ])
    _format_customer_block_sheet(ws, header_row=customer_header_row, data_start_row=customer_header_row + 1)
    ws.append([])
    ws.append(_receipt_goods_headers())
    goods_header_row = ws.max_row
    item_recs: list[list] = []
    for r in items:
        item_recs.append([
            r.get('inbound_date'), r.get('customer_name'), r.get('shop_no'), '',
            r.get('item_no'), r.get('item_name_cn'), r.get('material'),
            r.get('carton_count'), r.get('qty'),
            _num(r.get('unit_price')), _num(r.get('total_price')), _num(r.get('deposit_hint')),
            _num(r.get('cbm_at_load')),
            r.get('length_cm'), r.get('width_cm'), r.get('height_cm'),
        ])
    item_recs.sort(key=lambda r: (_group_key_text(r[1]).upper(), _group_key_text(r[2]).upper(), _group_key_text(r[4]).upper()))
    _append_grouped_rows_with_blanks(ws, item_recs, customer_idx=1, shop_idx=2)
    _style_sheet(ws, _receipt_col_widths())
    _format_receipt_goods_sheet(ws, header_row=goods_header_row)

    out = _export_dir() / f"container_{_sanitize_filename(head['container_no'])}_{today_str()}.xlsx"
    wb.save(out)
    return str(out)


def export_container_pdf(conn: Connection, container_id: int) -> str:
    head, items, customer_summary = container_manifest(conn, container_id)
    csum_sorted = _customer_rows_sync_like(customer_summary, master_customer_id=int(head.get('master_customer_id') or 0))
    out = _export_dir() / f"container_{_sanitize_filename(head['container_no'])}_{today_str()}.pdf"
    c = _new_pdf(out)
    width, height = A4
    y = height - 40
    c.drawString(40, y, f"柜号: {head['container_no']}  状态: {head['status']}  已用/容量: {_cbm_1(head['used_cbm'])}/{_cbm_1(head['capacity_cbm'])}")
    y -= 24
    c.drawString(40, y, 'PHONE')
    c.drawString(150, y, 'NAME')
    c.drawString(300, y, 'CBM')
    c.drawString(360, y, 'CTN')
    c.drawString(430, y, 'FREIGHT(CNY)')
    y -= 16
    for r in csum_sorted:
        if y < 60:
            c.showPage()
            y = height - 40
            c.setFont('STSong-Light', 11)
        c.drawString(40, y, str(r.get('customer_phone') or '')[:12])
        c.drawString(150, y, str(r.get('customer_name') or '')[:18])
        c.drawRightString(345, y, f"{_cbm_1(r.get('cbm_total')):.1f}")
        c.drawRightString(390, y, f"{int(r.get('ctns') or 0)}")
        c.drawRightString(530, y, f"{int(round(_num(r.get('freight_amount'))))}")
        y -= 14
    y -= 10
    c.drawString(40, y, 'ITEMS (customer/shop grouped)')
    y -= 16
    for r in sorted(items, key=lambda x: (_group_key_text(x.get('customer_name')).upper(), _group_key_text(x.get('shop_no')).upper(), _group_key_text(x.get('item_no')).upper())):
        if y < 60:
            c.showPage()
            y = height - 40
            c.setFont('STSong-Light', 11)
        c.drawString(40, y, str(r.get('inbound_date') or '')[:10])
        c.drawString(110, y, str(r.get('customer_name') or '')[:10])
        c.drawString(185, y, str(r.get('shop_no') or '')[:9])
        c.drawString(250, y, str(r.get('item_no') or '')[:10])
        c.drawString(325, y, str(r.get('item_name_cn') or '')[:10])
        c.drawRightString(390, y, f"{_cbm_1(r.get('cbm_at_load')):.1f}")
        y -= 14
    c.save()
    return str(out)
