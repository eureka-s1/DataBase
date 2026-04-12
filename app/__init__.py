from __future__ import annotations

from functools import wraps
from pathlib import Path
from uuid import uuid4
from datetime import datetime
import argparse
import re

import openpyxl

from flask import Flask, jsonify, redirect, render_template, request, session, url_for

from .config import BACKUP_DIR, DB_PATH, IMPORT_UPLOAD_DIR, SECRET_KEY
from .db import db_session, init_db
from .services.auth import authenticate, change_password
from .services.backup import backup_sqlite, list_backup_files, restore_sqlite_from_backup
from .services.containers import (
    add_item_to_container,
    container_manifest,
    confirm_container,
    container_usage,
    create_container,
    list_container_items,
    list_containers,
    remove_item_from_container,
    revoke_container,
    split_inbound_item_by_cartons,
    update_container_master_customer,
    update_container_no,
    update_item_cbm_at_load,
)
from .services.customers import (
    create_customer,
    find_customer_id_by_name,
    get_or_create_customer_id,
    list_customers,
    merge_customers,
    resolve_customer_id,
    update_customer_phone,
    upsert_alias,
)
from .services.finance import (
    add_payment,
    generate_statement,
    ledger,
    list_payments,
    list_statements,
    post_statement,
    post_statement_by_container,
    revoke_draft_statement_by_container,
    unpost_statement_by_container,
    unpost_statement,
)
from .services.importer import (
    import_inbound_excel,
    list_inbound_import_batches,
    parse_inbound_excel,
    rollback_inbound_import_batch,
)
from .services.inbound import create_inbound_item, delete_inbound_item, list_customer_items, list_inbound, update_inbound_item
from .services.pricing import upsert_price_rule
from .services.reports import (
    export_daily_inbound_excel,
    export_container_excel,
    export_container_pdf,
    export_inventory_excel,
    export_ledger_excel,
    export_statement_excel,
    export_statement_pdf,
)
from .services.ui_settings import (
    get_ui_settings,
    list_receipt_files,
    pick_receipt_file,
    pick_work_dir,
    set_monthly_auto_enabled,
    set_monthly_last_run_ym,
    set_work_dir,
)
from .services.file_sync import (
    ensure_sync_columns,
    list_outbound_sync_containers,
    list_receipt_sync_batches,
    monthly_create_sheet,
    sync_outbound_container,
    sync_outbound_container_to_customers,
    sync_outbound_container_to_manifest,
    sync_receipts_by_batch,
)
import scripts.import_historical_in_stock as hist_import


def _normalize_name_key(name: str) -> str:
    return re.sub(r"\s+", "", str(name or "")).upper().strip()


def _next_auto_customer_code(conn, prefix: str = "AUTOIMP") -> str:
    i = 1
    while True:
        code = f"{prefix}{i:05d}"
        row = conn.execute("SELECT 1 FROM customers WHERE customer_code=? LIMIT 1", (code,)).fetchone()
        if not row:
            return code
        i += 1


def _resolve_customer_id_by_main_name(conn, raw_name: str) -> int | None:
    key = _normalize_name_key(raw_name)
    if not key:
        return None
    row = conn.execute(
        "SELECT id FROM customers WHERE UPPER(REPLACE(name, ' ', ''))=? LIMIT 1",
        (key,),
    ).fetchone()
    return int(row["id"]) if row else None


def _ensure_customer_folder_and_empty_workbook(work_dir: Path, customer_name: str) -> Path:
    customer_dir = (work_dir / customer_name).resolve()
    customer_dir.mkdir(parents=True, exist_ok=True)
    now = datetime.now()
    ym = f"{now.year} {now.month}"
    xlsx_path = customer_dir / f"{customer_name} {now.year}收货清单.xlsx"
    if not xlsx_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ym
        wb.save(xlsx_path)
        return xlsx_path
    wb = openpyxl.load_workbook(xlsx_path)
    names = {str(x).strip() for x in wb.sheetnames}
    norm_names = {_normalize_name_key(x) for x in names}
    if _normalize_name_key(ym) not in norm_names and _normalize_name_key(f"{now.year} {now.month:02d}") not in norm_names:
        wb.create_sheet(title=ym)
    if wb.worksheets:
        wb.active = len(wb.worksheets) - 1
    wb.save(xlsx_path)
    return xlsx_path


def _collect_unresolved_customer_names(conn, file_path: Path) -> list[str]:
    parsed = parse_inbound_excel(file_path)
    names = []
    seen = set()
    for row in parsed.get("rows", []):
        n = str(row.get("customer_name") or "").strip()
        if not n:
            continue
        k = _normalize_name_key(n)
        if k in seen:
            continue
        seen.add(k)
        if resolve_customer_id(conn, n) is None:
            names.append(n)
    return names


def _apply_customer_review(
    conn,
    unresolved_names: list[str],
    decisions: list[dict],
    work_dir: Path,
) -> dict:
    by_key = {_normalize_name_key(x): x for x in unresolved_names}
    dmap = {}
    for d in decisions or []:
        raw_name = str((d or {}).get("raw_name") or "").strip()
        if not raw_name:
            continue
        dmap[_normalize_name_key(raw_name)] = d

    created_customers = []
    linked_aliases = []
    created_files = []
    missing = [x for x in unresolved_names if _normalize_name_key(x) not in dmap]
    if missing:
        raise ValueError(f"customer review decision missing: {', '.join(missing)}")

    for key, raw_name in by_key.items():
        d = dmap.get(key) or {}
        action = str(d.get("action") or "").strip().upper()
        if action not in ("ALIAS", "NEW"):
            raise ValueError(f"invalid action for {raw_name}")
        if action == "ALIAS":
            target_name = str(d.get("target_name") or "").strip()
            if not target_name:
                raise ValueError(f"target_name is required for alias action: {raw_name}")
            cid = resolve_customer_id(conn, target_name)
            if cid is None:
                raise ValueError(f"target customer not found: {target_name}")
            upsert_alias(
                conn,
                customer_id=int(cid),
                alias_name=raw_name,
                source="MANUAL",
                is_primary=0,
                is_active=1,
                remark="manual review during inbound import",
            )
            linked_aliases.append({"alias_name": raw_name, "target_name": target_name, "customer_id": int(cid)})
            continue

        cid = _resolve_customer_id_by_main_name(conn, raw_name)
        if cid is None:
            code = _next_auto_customer_code(conn)
            cid = create_customer(conn, customer_code=code, name=raw_name, default_price_per_m3=89.71)
            created_customers.append({"customer_id": int(cid), "name": raw_name})
        file_path = _ensure_customer_folder_and_empty_workbook(work_dir, raw_name)
        created_files.append({"customer_name": raw_name, "file_path": str(file_path)})

    return {
        "created_customers": created_customers,
        "linked_aliases": linked_aliases,
        "created_files": created_files,
    }


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login_page'))
        return fn(*args, **kwargs)

    return wrapper


def create_app() -> Flask:
    app = Flask(__name__)
    app.secret_key = SECRET_KEY

    @app.errorhandler(ValueError)
    def handle_value_error(err):
        if request.path.startswith('/ui/') or request.path in ('/', '/login'):
            return render_template('login.html', error=str(err)), 400
        return {'error': str(err)}, 400

    @app.route('/health', methods=['GET'])
    def health():
        return {'status': 'ok'}

    @app.route('/', methods=['GET'])
    @login_required
    def home():
        return redirect(url_for('ui_dashboard'))

    @app.route('/login', methods=['GET'])
    def login_page():
        return render_template('login.html')

    @app.route('/login', methods=['POST'])
    def login_submit():
        payload = request.form if request.form else request.get_json(force=True)
        username = (payload.get('username') or '').strip()
        password = (payload.get('password') or '').strip()
        with db_session() as conn:
            user = authenticate(conn, username, password)
        if not user:
            return render_template('login.html', error='用户名或密码错误'), 401

        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = user['role']
        if request.form:
            return redirect(url_for('ui_dashboard'))
        return user

    @app.route('/logout', methods=['POST'])
    @login_required
    def logout():
        session.clear()
        if request.form:
            return redirect(url_for('login_page'))
        return {'message': 'ok'}

    @app.route('/change-password', methods=['POST'])
    @login_required
    def change_password_api():
        payload = request.form if request.form else request.get_json(force=True)
        new_password = (payload.get('new_password') or '').strip()
        if len(new_password) < 6:
            return {'error': 'password too short'}, 400
        with db_session() as conn:
            change_password(conn, int(session['user_id']), new_password)
        return {'message': 'password changed'}

    @app.route('/init-db', methods=['POST'])
    def init_db_api():
        init_db()
        return {'message': 'database initialized'}

    @app.route('/customers', methods=['GET'])
    @login_required
    def customers_list():
        with db_session() as conn:
            return jsonify(list_customers(conn))

    @app.route('/customers', methods=['POST'])
    @login_required
    def customers_create():
        payload = request.get_json(force=True)
        required = ('customer_code', 'name')
        missing = [k for k in required if not payload.get(k)]
        if missing:
            return {'error': f'missing fields: {", ".join(missing)}'}, 400

        with db_session() as conn:
            customer_id = create_customer(
                conn,
                customer_code=payload['customer_code'],
                name=payload['name'],
                phone=payload.get('phone'),
                country=payload.get('country'),
                email=payload.get('email'),
                default_price_per_m3=float(payload.get('default_price_per_m3', 89.71)),
            )
        return {'id': customer_id}, 201

    @app.route('/customers/<int:customer_id>/phone', methods=['PUT'])
    @login_required
    def customers_update_phone_api(customer_id: int):
        payload = request.get_json(force=True)
        with db_session() as conn:
            update_customer_phone(conn, customer_id, payload.get('phone'))
        return {'message': 'ok'}

    @app.route('/customer-aliases', methods=['POST'])
    @login_required
    def alias_upsert():
        payload = request.get_json(force=True)
        customer_id = payload.get('customer_id')
        customer_name = (payload.get('customer_name') or '').strip()
        alias_name = payload.get('alias_name')
        if not alias_name:
            return {'error': 'alias_name is required'}, 400

        if customer_id:
            target_id = int(customer_id)
        elif customer_name:
            with db_session() as conn:
                target_id = find_customer_id_by_name(conn, customer_name)
            if target_id is None:
                return {'error': 'customer not found by customer_name'}, 400
        else:
            return {'error': 'customer_id or customer_name is required'}, 400

        with db_session() as conn:
            upsert_alias(
                conn,
                customer_id=target_id,
                alias_name=alias_name,
                source=payload.get('source', 'MANUAL'),
                is_primary=int(payload.get('is_primary', 0)),
                is_active=int(payload.get('is_active', 1)),
                remark=payload.get('remark'),
            )
        return {'message': 'ok'}

    @app.route('/customers/merge', methods=['POST'])
    @login_required
    def customers_merge_api():
        payload = request.get_json(force=True)
        source_id = int(payload['source_customer_id'])
        target_id = int(payload['target_customer_id'])
        with db_session() as conn:
            result = merge_customers(conn, source_id, target_id)
        return result

    @app.route('/customer-resolve', methods=['GET'])
    @login_required
    def customer_resolve():
        name = request.args.get('name', '').strip()
        if not name:
            return {'error': 'name is required'}, 400

        with db_session() as conn:
            customer_id = resolve_customer_id(conn, name)
        if customer_id is None:
            return {'matched': False}
        return {'matched': True, 'customer_id': customer_id}

    @app.route('/customer-items', methods=['GET'])
    @login_required
    def customer_items_api():
        customer_id_raw = (request.args.get('customer_id') or '').strip()
        customer_name = (request.args.get('customer_name') or '').strip()
        status = (request.args.get('status') or '').strip().upper() or None
        sort_by = (request.args.get('sort_by') or 'inbound_date').strip()
        sort_dir = (request.args.get('sort_dir') or 'desc').strip()

        if customer_id_raw:
            customer_id = int(customer_id_raw)
        elif customer_name:
            with db_session() as conn:
                customer_id = resolve_customer_id(conn, customer_name)
            if customer_id is None:
                return {'error': 'customer not found'}, 404
        else:
            return {'error': 'customer_id or customer_name is required'}, 400

        with db_session() as conn:
            rows = list_customer_items(conn, customer_id, status=status, sort_by=sort_by, sort_dir=sort_dir)
        return jsonify(rows)

    @app.route('/price-rules', methods=['POST'])
    @login_required
    def price_rules_upsert():
        payload = request.get_json(force=True)
        with db_session() as conn:
            upsert_price_rule(
                conn,
                customer_id=int(payload['customer_id']),
                effective_from=payload['effective_from'],
                effective_to=payload.get('effective_to'),
                price_per_m3=float(payload['price_per_m3']),
                currency=payload.get('currency', 'USD'),
                remark=payload.get('remark'),
            )
        return {'message': 'ok'}

    @app.route('/inbound-items', methods=['GET'])
    @login_required
    def inbound_list_api():
        inbound_date = request.args.get('inbound_date')
        only_in_stock = request.args.get('only_in_stock') == '1'
        batch_id_raw = (request.args.get('batch_id') or '').strip()
        import_batch_id = int(batch_id_raw) if batch_id_raw else None
        with db_session() as conn:
            return jsonify(
                list_inbound(
                    conn,
                    inbound_date=inbound_date,
                    only_in_stock=only_in_stock,
                    import_batch_id=import_batch_id,
                )
            )

    @app.route('/inbound-items', methods=['POST'])
    @login_required
    def inbound_create_api():
        payload = request.get_json(force=True)
        with db_session() as conn:
            item_id = create_inbound_item(conn, payload)
        return {'id': item_id}, 201

    @app.route('/inbound-items/<int:item_id>', methods=['PUT'])
    @login_required
    def inbound_update_api(item_id: int):
        payload = request.get_json(force=True)
        with db_session() as conn:
            update_inbound_item(conn, item_id, payload)
        return {'message': 'ok'}

    @app.route('/inbound-items/<int:item_id>', methods=['DELETE'])
    @login_required
    def inbound_delete_api(item_id: int):
        with db_session() as conn:
            deleted = delete_inbound_item(conn, item_id)
        if deleted == 0:
            return {'error': 'only IN_STOCK records can be deleted'}, 400
        return {'deleted': deleted}

    @app.route('/containers', methods=['POST'])
    @login_required
    def create_container_api():
        payload = request.get_json(force=True)
        with db_session() as conn:
            cid = create_container(conn, payload, user_id=int(session['user_id']))
        return {'id': cid}, 201

    @app.route('/containers', methods=['GET'])
    @login_required
    def containers_list_api():
        with db_session() as conn:
            return jsonify(list_containers(conn))

    @app.route('/containers/<int:container_id>', methods=['PUT'])
    @login_required
    def update_container_api(container_id: int):
        payload = request.get_json(force=True)
        with db_session() as conn:
            update_container_no(conn, container_id, payload.get('container_no', ''))
        return {'message': 'ok'}

    @app.route('/containers/<int:container_id>/master-customer', methods=['PUT'])
    @login_required
    def update_container_master_customer_api(container_id: int):
        payload = request.get_json(force=True)
        with db_session() as conn:
            update_container_master_customer(conn, container_id, payload.get('master_customer_id'))
        return {'message': 'ok'}

    @app.route('/containers/<int:container_id>/usage', methods=['GET'])
    @login_required
    def container_usage_api(container_id: int):
        with db_session() as conn:
            return container_usage(conn, container_id)

    @app.route('/containers/<int:container_id>/details', methods=['GET'])
    @login_required
    def container_details_api(container_id: int):
        with db_session() as conn:
            head, items, customer_summary = container_manifest(conn, container_id)
        return {'head': head, 'items': items, 'customer_summary': customer_summary}

    @app.route('/containers/<int:container_id>/items', methods=['POST'])
    @login_required
    def add_container_item_api(container_id: int):
        payload = request.get_json(force=True)
        with db_session() as conn:
            add_item_to_container(conn, container_id, int(payload['inbound_item_id']), payload.get('cbm_at_load'))
            usage = container_usage(conn, container_id)
        return {'message': 'ok', 'usage': usage}

    @app.route('/containers/<int:container_id>/items', methods=['GET'])
    @login_required
    def list_container_items_api(container_id: int):
        with db_session() as conn:
            rows = list_container_items(conn, container_id)
        return jsonify(rows)

    @app.route('/containers/<int:container_id>/items/<int:item_id>', methods=['DELETE'])
    @login_required
    def remove_container_item_api(container_id: int, item_id: int):
        with db_session() as conn:
            removed = remove_item_from_container(conn, container_id, item_id)
            usage = container_usage(conn, container_id)
        return {'removed': removed, 'usage': usage}

    @app.route('/containers/<int:container_id>/items/<int:item_id>', methods=['PUT'])
    @login_required
    def update_container_item_api(container_id: int, item_id: int):
        payload = request.get_json(force=True)
        cbm_at_load = float(payload['cbm_at_load'])
        with db_session() as conn:
            update_item_cbm_at_load(conn, container_id, item_id, cbm_at_load)
            usage = container_usage(conn, container_id)
        return {'message': 'ok', 'usage': usage}

    @app.route('/inbound-items/<int:item_id>/split-by-cartons', methods=['POST'])
    @login_required
    def split_inbound_item_api(item_id: int):
        payload = request.get_json(force=True)
        split_cartons = int(payload.get('split_cartons', 0))
        with db_session() as conn:
            result = split_inbound_item_by_cartons(
                conn,
                item_id,
                split_cartons,
                payload.get('length_cm'),
                payload.get('width_cm'),
                payload.get('height_cm'),
            )
        return result

    @app.route('/containers/<int:container_id>/confirm', methods=['POST'])
    @login_required
    def confirm_container_api(container_id: int):
        with db_session() as conn:
            confirm_container(conn, container_id)
        return {'message': 'confirmed'}

    @app.route('/containers/<int:container_id>/revoke', methods=['POST'])
    @login_required
    def revoke_container_api(container_id: int):
        with db_session() as conn:
            revoke_container(conn, container_id)
        return {'message': 'revoked'}

    @app.route('/containers/<int:container_id>/ship', methods=['POST'])
    @login_required
    def ship_container_api(container_id: int):
        with db_session() as conn:
            c = conn.execute('SELECT id, status FROM containers WHERE id=?', (container_id,)).fetchone()
            if not c:
                return {'error': 'container not found'}, 404
            if c['status'] != 'CONFIRMED':
                return {'error': 'only CONFIRMED container can be shipped'}, 400

            posted = conn.execute(
                "SELECT id FROM settlement_statements WHERE container_id=? AND status='POSTED' ORDER BY id DESC LIMIT 1",
                (container_id,),
            ).fetchone()
            if posted:
                return {'error': 'container already shipped (POSTED)'}, 400

            draft = conn.execute(
                "SELECT id FROM settlement_statements WHERE container_id=? AND status='DRAFT' ORDER BY id DESC LIMIT 1",
                (container_id,),
            ).fetchone()
            if not draft:
                generate_statement(conn, container_id=container_id, user_id=int(session['user_id']))
            statement_id = post_statement_by_container(conn, container_id)
        return {'message': 'shipped', 'container_id': container_id, 'statement_id': statement_id}

    @app.route('/payments', methods=['POST'])
    @login_required
    def add_payment_api():
        payload = request.get_json(force=True)
        with db_session() as conn:
            pid = add_payment(conn, payload, int(session['user_id']))
        return {'id': pid}, 201

    @app.route('/payments', methods=['GET'])
    @login_required
    def list_payments_api():
        limit = int(request.args.get('limit', 200))
        with db_session() as conn:
            rows = list_payments(conn, limit=limit)
        return jsonify(rows)

    @app.route('/settlements/generate', methods=['POST'])
    @login_required
    def generate_settlement_api():
        payload = request.get_json(force=True)
        with db_session() as conn:
            sid = generate_statement(
                conn,
                container_id=int(payload['container_id']),
                user_id=int(session['user_id']),
                statement_no=payload.get('statement_no'),
                statement_date=payload.get('statement_date'),
            )
        return {'statement_id': sid}, 201

    @app.route('/settlements/<int:statement_id>/post', methods=['POST'])
    @login_required
    def post_settlement_api(statement_id: int):
        with db_session() as conn:
            post_statement(conn, statement_id)
        return {'message': 'posted'}

    @app.route('/settlements/<int:statement_id>/unpost', methods=['POST'])
    @login_required
    def unpost_settlement_api(statement_id: int):
        with db_session() as conn:
            unpost_statement(conn, statement_id)
        return {'message': 'unposted'}

    @app.route('/settlements/container/<int:container_id>/post', methods=['POST'])
    @login_required
    def post_settlement_by_container_api(container_id: int):
        with db_session() as conn:
            statement_id = post_statement_by_container(conn, container_id)
        return {'message': 'posted', 'statement_id': statement_id}

    @app.route('/settlements/container/<int:container_id>/revoke', methods=['POST'])
    @login_required
    def revoke_settlement_by_container_api(container_id: int):
        with db_session() as conn:
            statement_id = revoke_draft_statement_by_container(conn, container_id)
        return {'message': 'revoked', 'statement_id': statement_id}

    @app.route('/settlements/container/<int:container_id>/unpost', methods=['POST'])
    @login_required
    def unpost_settlement_by_container_api(container_id: int):
        with db_session() as conn:
            statement_id = unpost_statement_by_container(conn, container_id)
        return {'message': 'unposted', 'statement_id': statement_id}

    @app.route('/settlements', methods=['GET'])
    @login_required
    def settlements_list_api():
        with db_session() as conn:
            rows = list_statements(conn)
        return jsonify(rows)

    @app.route('/ledger', methods=['GET'])
    @login_required
    def ledger_api():
        customer_id = request.args.get('customer_id')
        customer_name = (request.args.get('customer_name') or '').strip()
        with db_session() as conn:
            rows = ledger(
                conn,
                customer_id=int(customer_id) if customer_id else None,
                customer_name=customer_name or None,
            )
        return jsonify(rows)

    @app.route('/exports/daily-inbound', methods=['POST'])
    @login_required
    def export_daily_inbound_api():
        payload = request.get_json(force=True)
        with db_session() as conn:
            path = export_daily_inbound_excel(conn, payload.get('inbound_date'))
        return {'file_path': path}

    @app.route('/exports/inventory', methods=['POST'])
    @login_required
    def export_inventory_api():
        with db_session() as conn:
            path = export_inventory_excel(conn)
        return {'file_path': path}

    @app.route('/exports/ledger', methods=['POST'])
    @login_required
    def export_ledger_api():
        with db_session() as conn:
            path = export_ledger_excel(conn)
        return {'file_path': path}

    @app.route('/exports/statement/<int:statement_id>', methods=['POST'])
    @login_required
    def export_statement_api(statement_id: int):
        payload = request.get_json(force=True)
        fmt = (payload.get('format') or 'xlsx').lower()
        with db_session() as conn:
            if fmt == 'pdf':
                path = export_statement_pdf(conn, statement_id)
            else:
                path = export_statement_excel(conn, statement_id)
        return {'file_path': path}

    @app.route('/exports/statement/by-container/<int:container_id>', methods=['POST'])
    @login_required
    def export_statement_by_container_api(container_id: int):
        payload = request.get_json(force=True)
        fmt = (payload.get('format') or 'xlsx').lower()
        with db_session() as conn:
            row = conn.execute(
                '''
                SELECT id
                FROM settlement_statements
                WHERE container_id=? AND status IN ('DRAFT','POSTED')
                ORDER BY id DESC
                LIMIT 1
                ''',
                (container_id,),
            ).fetchone()
            if not row:
                return {'error': 'no settlement found for this container'}, 404
            statement_id = int(row['id'])
            if fmt == 'pdf':
                path = export_statement_pdf(conn, statement_id)
            else:
                path = export_statement_excel(conn, statement_id)
        return {'file_path': path, 'statement_id': statement_id}

    @app.route('/exports/container/<int:container_id>', methods=['POST'])
    @login_required
    def export_container_api(container_id: int):
        payload = request.get_json(force=True)
        fmt = (payload.get('format') or 'xlsx').lower()
        with db_session() as conn:
            if fmt == 'pdf':
                path = export_container_pdf(conn, container_id)
            else:
                path = export_container_excel(conn, container_id)
        return {'file_path': path}

    @app.route('/backup', methods=['POST'])
    @login_required
    def backup_api():
        with db_session() as conn:
            out = backup_sqlite(DB_PATH, BACKUP_DIR)
            conn.execute(
                'INSERT INTO backup_jobs(backup_time, backup_file, size_bytes, status, message) VALUES (datetime("now"), ?, ?, ?, ?)',
                (str(out), out.stat().st_size, 'SUCCESS', 'manual backup'),
            )
        return {'backup_file': str(out)}

    @app.route('/backups', methods=['GET'])
    @login_required
    def backups_api():
        return {"items": list_backup_files(BACKUP_DIR)}

    @app.route('/backups/restore', methods=['POST'])
    @login_required
    def restore_backup_api():
        payload = request.get_json(force=True) or {}
        raw_name = str(payload.get('file_name') or '').strip()
        file_name = Path(raw_name).name
        if not file_name:
            return {'error': 'file_name is required'}, 400
        backup_root = BACKUP_DIR.resolve()
        backup_file = (BACKUP_DIR / file_name).resolve()
        if str(backup_file.parent) != str(backup_root):
            return {'error': 'invalid backup file path'}, 400
        restore_sqlite_from_backup(DB_PATH, backup_file)
        with db_session() as conn:
            conn.execute(
                'INSERT INTO backup_jobs(backup_time, backup_file, size_bytes, status, message) VALUES (datetime("now"), ?, ?, ?, ?)',
                (str(backup_file), backup_file.stat().st_size, 'SUCCESS', 'manual restore'),
            )
        return {'ok': True, 'restored_from': str(backup_file)}

    @app.route('/import/inbound/preview', methods=['POST'])
    @login_required
    def import_preview_api():
        payload = request.get_json(force=True)
        path = Path(payload['file_path'])
        result = parse_inbound_excel(path)
        return {
            'header_row': result['header_row'],
            'sample_rows': result['rows'][:20],
            'field_mapping': result['field_mapping'],
        }

    @app.route('/import/inbound/upload', methods=['POST'])
    @login_required
    def import_upload_api():
        if 'file' not in request.files:
            return {'error': 'file is required'}, 400
        f = request.files['file']
        if not f or not f.filename:
            return {'error': 'empty filename'}, 400
        filename = Path(f.filename).name
        suffix = Path(filename).suffix.lower()
        if suffix not in ('.xlsx', '.xls'):
            return {'error': 'only .xlsx/.xls allowed'}, 400
        if '收货清单' not in filename:
            return {'error': '文件名必须包含“收货清单”'}, 400

        out = IMPORT_UPLOAD_DIR / f'{uuid4().hex[:10]}_{filename}'
        f.save(out)
        return {'file_path': str(out), 'filename': filename}

    @app.route('/import/inbound/workdir-files', methods=['GET'])
    @login_required
    def import_workdir_files_api():
        limit_raw = (request.args.get('limit') or '200').strip()
        limit = max(1, min(1000, int(limit_raw)))
        s = get_ui_settings()
        files = list_receipt_files(s.get('work_dir', ''), limit=limit)
        return {'work_dir': s.get('work_dir', ''), 'files': files}

    @app.route('/import/inbound/pick-file', methods=['POST'])
    @login_required
    def import_pick_file_api():
        payload = request.get_json(force=True) or {}
        return pick_receipt_file(payload.get('initial_dir'))

    @app.route('/import/inbound/execute', methods=['POST'])
    @login_required
    def import_execute_api():
        payload = request.get_json(force=True) or {}
        path = Path(payload['file_path'])
        inbound_date = payload.get('inbound_date')
        review_decisions = payload.get('customer_review') or []
        settings = get_ui_settings()
        work_dir = Path(settings.get('work_dir') or '').expanduser().resolve()
        if not work_dir.exists() or not work_dir.is_dir():
            return {'error': 'work directory not found'}, 400
        with db_session() as conn:
            unresolved = _collect_unresolved_customer_names(conn, path)
            if unresolved:
                if not review_decisions:
                    return {
                        'error': 'customer review required',
                        'error_code': 'CUSTOMER_REVIEW_REQUIRED',
                        'unknown_customers': unresolved,
                    }, 409
                review_result = _apply_customer_review(
                    conn,
                    unresolved_names=unresolved,
                    decisions=review_decisions,
                    work_dir=work_dir,
                )
                unresolved_after = _collect_unresolved_customer_names(conn, path)
                if unresolved_after:
                    return {
                        'error': f'unresolved customers after review: {", ".join(unresolved_after)}',
                    }, 400
            else:
                review_result = {
                    "created_customers": [],
                    "linked_aliases": [],
                    "created_files": [],
                }
            result = import_inbound_excel(
                conn,
                path=path,
                inbound_date=inbound_date,
                created_by=int(session['user_id']),
                dry_run=False,
                auto_create_customers=False,
            )
        result['customer_review'] = review_result
        return result

    @app.route('/import/inbound/manual-item', methods=['POST'])
    @login_required
    def import_manual_single_item_api():
        payload = request.get_json(force=True) or {}
        customer_name = str(payload.get('customer_name') or '').strip()
        if not customer_name:
            return {'error': 'customer_name is required'}, 400
        item_name = str(payload.get('item_name_cn') or '').strip()
        if not item_name:
            return {'error': 'item_name_cn is required'}, 400

        with db_session() as conn:
            customer_id, customer_created = get_or_create_customer_id(conn, customer_name)
            wh = conn.execute("SELECT id FROM warehouses ORDER BY id LIMIT 1").fetchone()
            if not wh:
                return {'error': 'warehouse not found'}, 400
            warehouse_id = int(wh['id'])

            inbound_no = str(payload.get('inbound_no') or '').strip()
            if not inbound_no:
                inbound_no = f"MAN-{datetime.now().strftime('%Y%m%d')}-{uuid4().hex[:10].upper()}"
            while conn.execute("SELECT 1 FROM inbound_items WHERE inbound_no=? LIMIT 1", (inbound_no,)).fetchone():
                inbound_no = f"MAN-{datetime.now().strftime('%Y%m%d')}-{uuid4().hex[:10].upper()}"

            inbound_date = str(payload.get('inbound_date') or '').strip() or datetime.now().strftime('%Y-%m-%d')

            item_payload = {
                'inbound_no': inbound_no,
                'import_batch_id': None,
                'customer_id': customer_id,
                'warehouse_id': warehouse_id,
                'inbound_date': inbound_date,
                'shop_no': payload.get('shop_no'),
                'position_or_tel': payload.get('position_or_tel'),
                'item_no': payload.get('item_no'),
                'item_name_cn': item_name,
                'material': payload.get('material'),
                'carton_count': payload.get('carton_count'),
                'qty': payload.get('qty'),
                'unit_price': payload.get('unit_price'),
                'total_price': payload.get('total_price'),
                'deposit_hint': payload.get('deposit_hint'),
                'length_cm': payload.get('length_cm'),
                'width_cm': payload.get('width_cm'),
                'height_cm': payload.get('height_cm'),
                'cbm_calculated': payload.get('cbm_calculated'),
                'cbm_override': payload.get('cbm_override'),
                'status': 'IN_STOCK',
                'remark': payload.get('remark') or 'MANUAL_SINGLE_IMPORT',
            }
            item_id = create_inbound_item(conn, item_payload)
        return {
            'id': int(item_id),
            'customer_id': int(customer_id),
            'customer_auto_created': bool(customer_created),
            'inbound_no': inbound_no,
        }, 201

    @app.route('/import/historical-in-stock/execute', methods=['POST'])
    @login_required
    def import_historical_in_stock_execute_api():
        payload = request.get_json(force=True) or {}
        settings = get_ui_settings()
        data_root = Path(payload.get('data_root') or settings.get('work_dir') or '').resolve()
        if not data_root.exists() or not data_root.is_dir():
            return {'error': 'work directory not found'}, 400
        dry_run = bool(payload.get('dry_run', True))
        min_file_year = int(payload.get('min_file_year') or 0)
        inbound_date = str(payload.get('inbound_date') or '').strip()
        args = argparse.Namespace(
            data_root=data_root,
            customer=None,
            limit=0,
            inbound_date=inbound_date,
            min_file_year=min_file_year,
            dry_run=dry_run,
            apply=not dry_run,
            verbose=False,
        )
        report = hist_import.run_import(args)
        return report

    @app.route('/import/inbound/rollback/<int:batch_id>', methods=['POST'])
    @login_required
    def import_rollback_api(batch_id: int):
        with db_session() as conn:
            result = rollback_inbound_import_batch(conn, batch_id)
        return result

    @app.route('/import/inbound/batches', methods=['GET'])
    @login_required
    def import_batches_api():
        batch_id_raw = (request.args.get('batch_id') or '').strip()
        inbound_date = (request.args.get('inbound_date') or '').strip() or None
        limit_raw = (request.args.get('limit') or '50').strip()
        limit = max(1, min(500, int(limit_raw)))
        with db_session() as conn:
            rows = list_inbound_import_batches(
                conn,
                limit=limit,
                batch_id=int(batch_id_raw) if batch_id_raw else None,
                inbound_date=inbound_date,
            )
        return jsonify(rows)

    @app.route('/dashboard/summary', methods=['GET'])
    @login_required
    def dashboard_summary_api():
        with db_session() as conn:
            customer_count = int(conn.execute('SELECT COUNT(*) AS c FROM customers WHERE is_active=1').fetchone()['c'])
            inbound_count = int(conn.execute("SELECT COUNT(*) AS c FROM inbound_items WHERE status='IN_STOCK'").fetchone()['c'])
            container_count = int(conn.execute('SELECT COUNT(*) AS c FROM containers').fetchone()['c'])
            statement_count = int(conn.execute('SELECT COUNT(*) AS c FROM settlement_statements').fetchone()['c'])
        return {
            'customer_count': customer_count,
            'inbound_count': inbound_count,
            'container_count': container_count,
            'statement_count': statement_count,
        }

    @app.route('/settings', methods=['GET'])
    @login_required
    def settings_api():
        return get_ui_settings()

    @app.route('/settings/work-dir', methods=['PUT'])
    @login_required
    def settings_work_dir_api():
        payload = request.get_json(force=True)
        return set_work_dir(payload.get('work_dir'))

    @app.route('/settings/work-dir/pick', methods=['POST'])
    @login_required
    def settings_work_dir_pick_api():
        payload = request.get_json(force=True) or {}
        return pick_work_dir(payload.get('initial_dir'))

    @app.route('/settings/clear-db', methods=['POST'])
    @login_required
    def settings_clear_db_api():
        if str(session.get('role') or '') != 'admin':
            return {'error': 'admin required'}, 403
        payload = request.get_json(force=True) or {}
        admin_password = str(payload.get('admin_password') or '').strip()
        if not admin_password:
            return {'error': 'admin_password is required'}, 400

        with db_session() as conn:
            # Require current admin account password confirmation.
            user = authenticate(conn, str(session.get('username') or ''), admin_password)
            if not user or str(user.get('role') or '') != 'admin':
                return {'error': '管理员密码错误'}, 400

            tables = [
                'audit_logs',
                'payment_allocations',
                'settlement_lines',
                'settlement_statements',
                'payment_transactions',
                'container_items',
                'inbound_import_rows',
                'inbound_items',
                'import_batches',
                'customer_price_rules',
                'customer_aliases',
                'customers',
                'containers',
                'export_jobs',
                'backup_jobs',
            ]
            conn.execute('PRAGMA foreign_keys = OFF')
            try:
                for t in tables:
                    conn.execute(f'DELETE FROM {t}')
                seq_exists = conn.execute(
                    "SELECT 1 FROM sqlite_master WHERE type='table' AND name='sqlite_sequence' LIMIT 1"
                ).fetchone()
                if seq_exists:
                    conn.execute(
                        "DELETE FROM sqlite_sequence WHERE name IN (" + ",".join("?" for _ in tables) + ")",
                        tables,
                    )
                # Keep at least one warehouse for subsequent inbound writes.
                conn.execute('DELETE FROM warehouses')
                conn.execute(
                    "INSERT INTO warehouses(name, location, is_active) VALUES (?, ?, 1)",
                    ("Main Warehouse", "Default"),
                )
            finally:
                conn.execute('PRAGMA foreign_keys = ON')

        return {'ok': True, 'message': 'database business data cleared'}

    @app.route('/sync/receipts/batches', methods=['GET'])
    @login_required
    def sync_receipt_batches_api():
        limit = max(1, min(1000, int((request.args.get('limit') or '200').strip())))
        with db_session() as conn:
            ensure_sync_columns(conn)
            rows = list_receipt_sync_batches(conn, limit=limit)
        return jsonify(rows)

    @app.route('/sync/receipts/batch/<int:batch_id>', methods=['POST'])
    @login_required
    def sync_receipt_batch_execute_api(batch_id: int):
        s = get_ui_settings()
        work_dir = Path(s.get('work_dir') or '').expanduser().resolve()
        if not work_dir.exists() or not work_dir.is_dir():
            return {'error': 'work directory not found'}, 400
        with db_session() as conn:
            ensure_sync_columns(conn)
            result = sync_receipts_by_batch(conn, batch_id, work_dir)
        return result

    @app.route('/sync/outbound/containers', methods=['GET'])
    @login_required
    def sync_outbound_containers_api():
        limit = max(1, min(1000, int((request.args.get('limit') or '300').strip())))
        with db_session() as conn:
            ensure_sync_columns(conn)
            rows = list_outbound_sync_containers(conn, limit=limit)
        return jsonify(rows)

    @app.route('/sync/outbound/container/<int:container_id>', methods=['POST'])
    @login_required
    def sync_outbound_container_execute_api(container_id: int):
        s = get_ui_settings()
        work_dir = Path(s.get('work_dir') or '').expanduser().resolve()
        if not work_dir.exists() or not work_dir.is_dir():
            return {'error': 'work directory not found'}, 400
        with db_session() as conn:
            ensure_sync_columns(conn)
            result = sync_outbound_container(conn, container_id, work_dir)
        return result

    @app.route('/sync/outbound/container/<int:container_id>/to-customers', methods=['POST'])
    @login_required
    def sync_outbound_container_to_customers_api(container_id: int):
        s = get_ui_settings()
        work_dir = Path(s.get('work_dir') or '').expanduser().resolve()
        if not work_dir.exists() or not work_dir.is_dir():
            return {'error': 'work directory not found'}, 400
        with db_session() as conn:
            ensure_sync_columns(conn)
            result = sync_outbound_container_to_customers(conn, container_id, work_dir)
        return result

    @app.route('/sync/outbound/container/<int:container_id>/to-manifest', methods=['POST'])
    @login_required
    def sync_outbound_container_to_manifest_api(container_id: int):
        payload = request.get_json(force=True) or {}
        allow_create = bool(payload.get('allow_create', False))
        s = get_ui_settings()
        work_dir = Path(s.get('work_dir') or '').expanduser().resolve()
        if not work_dir.exists() or not work_dir.is_dir():
            return {'error': 'work directory not found'}, 400
        with db_session() as conn:
            ensure_sync_columns(conn)
            result = sync_outbound_container_to_manifest(conn, container_id, work_dir, allow_create=allow_create)
        return result

    @app.route('/sync/monthly/settings', methods=['PUT'])
    @login_required
    def sync_monthly_settings_api():
        payload = request.get_json(force=True) or {}
        enabled = bool(payload.get('monthly_auto_enabled', True))
        return set_monthly_auto_enabled(enabled)

    @app.route('/sync/monthly/execute', methods=['POST'])
    @login_required
    def sync_monthly_execute_api():
        payload = request.get_json(force=True) or {}
        year = int(payload.get('year') or 0)
        month = int(payload.get('month') or 0)
        if year < 2000 or year > 2100 or month < 1 or month > 12:
            return {'error': 'invalid year/month'}, 400
        s = get_ui_settings()
        work_dir = Path(s.get('work_dir') or '').expanduser().resolve()
        if not work_dir.exists() or not work_dir.is_dir():
            return {'error': 'work directory not found'}, 400
        result = monthly_create_sheet(work_dir, year, month)
        set_monthly_last_run_ym(result.ym)
        return {
            'ym': result.ym,
            'files_scanned': result.files_scanned,
            'files_updated': result.files_updated,
            'errors': result.errors,
        }

    @app.route('/sync/monthly/auto-status', methods=['GET'])
    @login_required
    def sync_monthly_auto_status_api():
        s = get_ui_settings()
        from datetime import datetime

        now = datetime.now()
        ym = f'{now.year:04d} {now.month}'
        should_prompt = bool(s.get('monthly_auto_enabled')) and now.day == 1 and str(s.get('monthly_last_run_ym') or '') != ym
        return {
            'today': now.strftime('%Y-%m-%d'),
            'ym': ym,
            'monthly_auto_enabled': bool(s.get('monthly_auto_enabled')),
            'monthly_last_run_ym': str(s.get('monthly_last_run_ym') or ''),
            'should_prompt': should_prompt,
        }

    @app.route('/ui/dashboard', methods=['GET'])
    @login_required
    def ui_dashboard():
        with db_session() as conn:
            customers = list_customers(conn)
            inbound = list_inbound(conn, only_in_stock=True)
            containers = list_containers(conn)
            ledger_rows = ledger(conn)
            statements = list_statements(conn, limit=30)
        return render_template(
            'dashboard.html',
            user=session,
            customers=customers,
            inbound=inbound[:50],
            containers=containers[:30],
            ledger_rows=ledger_rows,
            statements=statements,
        )

    return app
